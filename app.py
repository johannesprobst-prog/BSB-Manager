import io
import os
import sqlite3
import zipfile
from datetime import date, datetime, timedelta
from PIL import Image, ImageDraw
import numpy as np
import pandas as pd
import streamlit as st
from streamlit_drawable_canvas import st_canvas
from streamlit_image_coordinates import streamlit_image_coordinates

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import qrcode

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image as RLImage,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

import fitz  # PyMuPDF für PDF-Pläne

DB_FILE = "brandschutz.db"
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def get_db():
  conn = sqlite3.connect(DB_FILE)
  conn.row_factory = sqlite3.Row
  return conn


def init_db():
  conn = get_db()
  cursor = conn.cursor()
  cursor.executescript("""
    CREATE TABLE IF NOT EXISTS properties (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        address TEXT NOT NULL,
        fire_safety_officer TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS floor_plans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        property_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        image_path TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (property_id) REFERENCES properties(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS fire_facilities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        property_id INTEGER NOT NULL,
        floor_plan_id INTEGER,
        pin_x REAL,
        pin_y REAL,
        category TEXT NOT NULL,
        identifier TEXT NOT NULL,
        device_type TEXT,
        location_desc TEXT NOT NULL,
        norm_ref TEXT,
        last_inspection TEXT NOT NULL DEFAULT 'DAUERHAFT',
        next_inspection TEXT NOT NULL DEFAULT 'DAUERHAFT',
        inspection_interval_months INTEGER DEFAULT 24,
        inspector_company TEXT,
        has_defect BOOLEAN DEFAULT 0,
        defect_description TEXT,
        defect_severity TEXT,
        defect_due_date DATE,
        defect_photo_path TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (property_id) REFERENCES properties(id) ON DELETE CASCADE,
        FOREIGN KEY (floor_plan_id) REFERENCES floor_plans(id) ON DELETE SET NULL
    );

    CREATE TABLE IF NOT EXISTS checklist_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        facility_category TEXT NOT NULL,
        trvb_ref TEXT NOT NULL,
        interval TEXT NOT NULL,
        check_item TEXT NOT NULL,
        instruction TEXT
    );

    CREATE TABLE IF NOT EXISTS inspection_runs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        property_id INTEGER NOT NULL,
        inspector_name TEXT NOT NULL,
        inspection_date DATE NOT NULL,
        interval_scope TEXT NOT NULL,
        status TEXT DEFAULT 'IN_BEGEHUNG',
        sig_bsb_path TEXT,
        sig_mgmt_path TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (property_id) REFERENCES properties(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS inspection_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        inspection_run_id INTEGER NOT NULL,
        checklist_template_id INTEGER NOT NULL,
        result_status TEXT NOT NULL,
        remark TEXT,
        FOREIGN KEY (inspection_run_id) REFERENCES inspection_runs(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS defects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        property_id INTEGER NOT NULL,
        inspection_run_id INTEGER,
        floor_plan_id INTEGER,
        facility_category TEXT,
        pin_x REAL,
        pin_y REAL,
        title TEXT NOT NULL,
        description TEXT,
        severity TEXT NOT NULL,
        status TEXT DEFAULT 'OFFEN',
        due_date DATE,
        image_path TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS journal_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        property_id INTEGER NOT NULL,
        event_date DATE NOT NULL,
        event_time TEXT,
        event_type TEXT NOT NULL,
        sub_type TEXT,
        title TEXT NOT NULL,
        details TEXT,
        detector_group TEXT,
        fire_brigade_deployed BOOLEAN DEFAULT 0,
        participants_count INTEGER,
        duration_minutes INTEGER,
        instructor_name TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (property_id) REFERENCES properties(id) ON DELETE CASCADE
    );
    """)

  cursor.execute("SELECT COUNT(*) FROM checklist_templates")
  if cursor.fetchone()[0] == 0:
    seed_items = [
        (
            "FLUECHTWEGE",
            "TRVB 117 O",
            "MONATLICH",
            "Fluchtwege frei von unzulässigen Brandlasten",
            "Gänge, Notausgänge und Treppenhäuser kontrollieren.",
        ),
        (
            "FLUECHTWEGE",
            "TRVB 117 O",
            "MONATLICH",
            "Notausgänge leicht öffenbar und unversperrt",
            "Panikbeschläge und Verriegelungen testen.",
        ),
        (
            "BMA",
            "TRVB 123 S",
            "MONATLICH",
            "BMZ im Normalbetrieb & Meldertest",
            "Probeauslösung eines automatischen Melders.",
        ),
        (
            "TUEREN",
            "TRVB 148 S",
            "MONATLICH",
            "Brandschutztüren schließen einwandfrei",
            "Selbstschließung und Dichtungen prüfen.",
        ),
        (
            "LOESCHER",
            "TRVB 124 S",
            "MONATLICH",
            "Feuerlöscher zugänglich & Prüffrist gültig",
            "Plombe intakt, Manometer grün, Prüfplakette gültig.",
        ),
        (
            "RWA",
            "TRVB 125 S",
            "MONATLICH",
            "Auslöseeinrichtungen & Manometer geprüft",
            "Optische Prüfung der Handauslöser und Druckbehälter.",
        ),
        (
            "NOTLICHT",
            "TRVB 117 O",
            "MONATLICH",
            "Sicherheitsbeleuchtung Funktionstest",
            "Optische Prüfung der Betriebsbereitschaft.",
        ),
    ]
    cursor.executemany(
        "INSERT INTO checklist_templates (facility_category, trvb_ref,"
        " interval, check_item, instruction) VALUES (?, ?, ?, ?, ?)",
        seed_items,
    )
    cursor.execute(
        "INSERT INTO properties (name, address, fire_safety_officer) VALUES"
        " (?, ?, ?)",
        (
            "Musterbetrieb Werk 1",
            "Musterstraße 1, 5134 Schwand",
            "Johannes Probst",
        ),
    )
  conn.commit()
  conn.close()


init_db()


def draw_color_dot(
    draw, x, y, fill_color, outline_color="white", label="", radius=13
):
  draw.ellipse(
      (x - radius, y - radius, x + radius, y + radius),
      fill=fill_color,
      outline=outline_color,
      width=2,
  )
  if label:
    draw.text((x - 4, y - 6), str(label), fill="white")


# --- PDF GENERATOR FÜR BSB-JAHRESBERICHT ---
def generate_annual_report_pdf(property_id, report_year, bsb_summary_text):
  conn = get_db()
  prop = conn.execute(
      "SELECT * FROM properties WHERE id = ?", (property_id,)
  ).fetchone()

  runs = conn.execute(
      """
        SELECT * FROM inspection_runs 
        WHERE property_id = ? AND strftime('%Y', inspection_date) = ?
        ORDER BY inspection_date ASC
    """,
      (property_id, str(report_year)),
  ).fetchall()

  events = conn.execute(
      """
        SELECT * FROM journal_events 
        WHERE property_id = ? AND strftime('%Y', event_date) = ?
        ORDER BY event_date ASC
    """,
      (property_id, str(report_year)),
  ).fetchall()

  facilities = conn.execute(
      """
        SELECT * FROM fire_facilities 
        WHERE property_id = ?
    """,
      (property_id,),
  ).fetchall()
  conn.close()

  alarms = [e for e in events if e["event_type"] == "ALARM"]
  false_alarms = [
      e for e in alarms if e["sub_type"] in ["TAEUSCHUNG", "FEHLALARM"]
  ]
  real_alarms = [e for e in alarms if e["sub_type"] == "ECHTALARM"]
  fw_deployments = [e for e in alarms if e["fire_brigade_deployed"]]

  exercises = [e for e in events if e["event_type"] == "UEBUNG"]
  trainings = [e for e in events if e["event_type"] == "SCHULUNG"]
  heissarbeiten = [e for e in events if e["event_type"] == "HEISSARBEIT"]
  defective_facs = [f for f in facilities if f["has_defect"]]

  buffer = io.BytesIO()
  doc = SimpleDocTemplate(
      buffer,
      pagesize=A4,
      leftMargin=15 * mm,
      rightMargin=15 * mm,
      topMargin=15 * mm,
      bottomMargin=15 * mm,
  )

  styles = getSampleStyleSheet()
  title_style = ParagraphStyle(
      "AR_Title",
      parent=styles["Heading1"],
      fontSize=15,
      leading=18,
      textColor=colors.HexColor("#0F172A"),
      spaceAfter=2,
  )
  sub_style = ParagraphStyle(
      "AR_Sub",
      parent=styles["Normal"],
      fontSize=8.5,
      leading=11,
      textColor=colors.HexColor("#475569"),
  )
  h2_style = ParagraphStyle(
      "AR_H2",
      parent=styles["Heading2"],
      fontSize=10.5,
      leading=13,
      textColor=colors.HexColor("#1E293B"),
      spaceBefore=5,
      spaceAfter=3,
  )
  th_style = ParagraphStyle(
      "AR_TH",
      parent=styles["Normal"],
      fontSize=8,
      leading=10,
      fontName="Helvetica-Bold",
      textColor=colors.white,
  )
  td_style = ParagraphStyle(
      "AR_TD",
      parent=styles["Normal"],
      fontSize=7.5,
      leading=9.5,
      textColor=colors.HexColor("#1E293B"),
  )

  elements = [
      Paragraph("JAHRESBERICHT DES BRANDSCHUTZBEAUFTRAGTEN", title_style),
      Paragraph(
          f"Berichtsjahr {report_year} gem. TRVB 117 O für die"
          " Geschäftsleitung / Betriebsführung",
          sub_style,
      ),
      Spacer(1, 3 * mm),
  ]

  meta_data = [
      [
          Paragraph("<b>Liegenschaft / Betrieb:</b>", td_style),
          Paragraph(f"{prop['name']}<br/>{prop['address']}", td_style),
          Paragraph("<b>Berichtszeitraum:</b>", td_style),
          Paragraph(
              f"01.01.{report_year} – 31.12.{report_year}<br/>Druck:"
              f" {date.today()}",
              td_style,
          ),
      ],
      [
          Paragraph("<b>Brandschutzbeauftragter:</b>", td_style),
          Paragraph(f"{prop['fire_safety_officer']}", td_style),
          Paragraph("<b>Status Brandschutzbuch:</b>", td_style),
          Paragraph("<b><font color='#16A34A'>Vollständig geführt</font></b>", td_style),
      ],
  ]
  t_meta = Table(meta_data, colWidths=[38 * mm, 62 * mm, 35 * mm, 45 * mm])
  t_meta.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
          ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
          ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
          ("TOPPADDING", (0, 0), (-1, -1), 3),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
      ])
  )
  elements.append(t_meta)
  elements.append(Spacer(1, 4 * mm))

  elements.append(
      Paragraph("1. Brandschutz-Kennzahlen im Berichtsjahr", h2_style)
  )
  kpi_data = [
      [
          Paragraph("<b>Kennzahl</b>", th_style),
          Paragraph("<b>Anzahl / Wert</b>", th_style),
          Paragraph("<b>Bemerkung / TRVB-Vorgabe</b>", th_style),
      ],
      [
          Paragraph("Durchgeführte BSB-Eigenkontrollen", td_style),
          Paragraph(f"<b>{len(runs)} Begehung(en)</b>", td_style),
          Paragraph(
              "Monatlich / Quartal gem. TRVB 117 O ordnungsgemäß erfüllt",
              td_style,
          ),
      ],
      [
          Paragraph("Mitarbeiter-Unterweisungen & Schulungen", td_style),
          Paragraph(
              f"<b>{len(trainings)} Schulung(en)</b>"
              f" ({sum(t['participants_count'] or 0 for t in trainings)} MA)",
              td_style,
          ),
          Paragraph(
              "Mind. 1x jährlich gefordert gem. ASchG / TRVB 117 O", td_style
          ),
      ],
      [
          Paragraph("Evakuierungs- & Räumungsübungen", td_style),
          Paragraph(f"<b>{len(exercises)} Übung(en)</b>", td_style),
          Paragraph(
              "Regelmäßige Räumungsübung mit Belegschaft durchgeführt", td_style
          ),
      ],
      [
          Paragraph("Brandmeldealarme gesamt", td_style),
          Paragraph(f"<b>{len(alarms)} Alarm(e)</b>", td_style),
          Paragraph(
              f"{len(real_alarms)} Realbrand / {len(false_alarms)}"
              f" Täuschungs-/Fehlalarme ({len(fw_deployments)} FW-Einsätze)",
              td_style,
          ),
      ],
      [
          Paragraph("Freigabescheine für Heißarbeiten", td_style),
          Paragraph(f"<b>{len(heissarbeiten)} Arbeiten</b>", td_style),
          Paragraph("Schweiß-/Schneidarbeiten mit Brandwache", td_style),
      ],
      [
          Paragraph("Aktuell offene Mängel an Einrichtungen", td_style),
          Paragraph(
              f"<b><font color='{'#DC2626' if defective_facs else '#16A34A'}'>{len(defective_facs)} Mängel</font></b>",
              td_style,
          ),
          Paragraph(
              "Siehe Mängelverwaltung & Instandsetzungsaufträge", td_style
          ),
      ],
  ]
  t_kpi = Table(kpi_data, colWidths=[65 * mm, 35 * mm, 80 * mm])
  t_kpi.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
          ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
          ("TOPPADDING", (0, 0), (-1, -1), 3),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
          ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
      ])
  )
  elements.append(t_kpi)
  elements.append(Spacer(1, 4 * mm))

  elements.append(
      Paragraph("2. Bestand der Brandschutzeinrichtungen", h2_style)
  )
  cats_count = {}
  for f in facilities:
    c = f["category"]
    cats_count[c] = cats_count.get(c, 0) + 1

  cat_rows = [
      f"<b>{c}</b>: {cnt} Stk." for c, cnt in sorted(cats_count.items())
  ]
  cat_str = " | ".join(cat_rows) if cat_rows else "Keine Geräte erfasst."
  elements.append(
      Paragraph(
          f"Gesamtanzahl erfasster Einrichtungen: <b>{len(facilities)}</b>"
          f" ({cat_str})",
          td_style,
      )
  )
  elements.append(Spacer(1, 4 * mm))

  elements.append(
      Paragraph(
          "3. Gesamtbeurteilung des baulichen & organisatorischen"
          " Brandschutzes",
          h2_style,
      )
  )
  summary_p = bsb_summary_text if bsb_summary_text else (
      "Der betriebliche Brandschutz entspricht im Berichtszeitraum den"
      " gesetzlichen und behördlichen Anforderungen. Die erforderlichen"
      " Eigenkontrollen wurden lückenlos durchgeführt. Die festgestellten"
      " Mängel wurden bzw. werden zeitgerecht im Rahmen des"
      " Instandsetzungsmanagements abgearbeitet."
  )
  t_sum = Table(
      [[Paragraph(summary_p, td_style)]],
      colWidths=[180 * mm],
  )
  t_sum.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F1F5F9")),
          ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
          ("TOPPADDING", (0, 0), (-1, -1), 5),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
      ])
  )
  elements.append(t_sum)
  elements.append(Spacer(1, 8 * mm))

  sign_data = [[
      Paragraph(
          "____________________________________________<br/>Datum, Unterschrift"
          " Brandschutzbeauftragter",
          td_style,
      ),
      Paragraph(
          "____________________________________________<br/>Kenntnisnahme"
          " Betriebsleitung / Eigentümer",
          td_style,
      ),
  ]]
  s_table = Table(sign_data, colWidths=[90 * mm, 90 * mm])
  s_table.setStyle(
      TableStyle([
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("LEFTPADDING", (0, 0), (-1, -1), 0),
      ])
  )
  elements.append(KeepTogether([s_table]))

  doc.build(elements)
  buffer.seek(0)
  return buffer.getvalue()


# --- QR-CODE ETIKETTENBOGEN GENERATOR ---
def generate_qr_labels_pdf(property_id, selected_ids=None):
  conn = get_db()
  prop = conn.execute(
      "SELECT * FROM properties WHERE id = ?", (property_id,)
  ).fetchone()

  query = """
        SELECT f.*, fp.name as plan_name 
        FROM fire_facilities f
        LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
        WHERE f.property_id = ?
    """
  params = [property_id]
  if selected_ids:
    placeholders = ",".join("?" for _ in selected_ids)
    query += f" AND f.id IN ({placeholders})"
    params.extend(selected_ids)

  query += " ORDER BY f.category, f.identifier ASC"
  facilities = conn.execute(query, params).fetchall()
  conn.close()

  buffer = io.BytesIO()
  doc = SimpleDocTemplate(
      buffer,
      pagesize=A4,
      leftMargin=10 * mm,
      rightMargin=10 * mm,
      topMargin=12 * mm,
      bottomMargin=12 * mm,
  )

  styles = getSampleStyleSheet()
  label_id_style = ParagraphStyle(
      "LID",
      parent=styles["Normal"],
      fontSize=11,
      leading=13,
      fontName="Helvetica-Bold",
      textColor=colors.HexColor("#0F172A"),
  )
  label_text_style = ParagraphStyle(
      "LTXT",
      parent=styles["Normal"],
      fontSize=7.5,
      leading=9.5,
      textColor=colors.HexColor("#334155"),
  )

  elements = []
  label_cells = []

  for f in facilities:
    qr_payload = f"BSB-GERAET|ID:{f['id']}|IDENT:{f['identifier']}|CAT:{f['category']}|OBJ:{prop['name']}"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=4,
        border=1,
    )
    qr.add_data(qr_payload)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")

    qr_byte_arr = io.BytesIO()
    qr_img.save(qr_byte_arr, format="PNG")
    qr_byte_arr.seek(0)
    rl_qr = RLImage(qr_byte_arr, width=22 * mm, height=22 * mm)

    frist_str = (
        f"Prüfung: {f['next_inspection']}"
        if f["next_inspection"] not in ["DAUERHAFT", "None", ""]
        else "Dauerhaft gültig"
    )

    text_info = f"""
        <b>{f['identifier']}</b> ({f['category']})<br/>
        {f['device_type'] or ''}<br/>
        <font color='#64748B'>{f['location_desc']}</font><br/>
        <b>{frist_str}</b>
        """

    inner_table_data = [[rl_qr, Paragraph(text_info, label_text_style)]]
    inner_table = Table(inner_table_data, colWidths=[24 * mm, 36 * mm])
    inner_table.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )

    label_cells.append(inner_table)

  rows = []
  current_row = []
  for cell in label_cells:
    current_row.append(cell)
    if len(current_row) == 3:
      rows.append(current_row)
      current_row = []

  if current_row:
    while len(current_row) < 3:
      current_row.append(Paragraph("", label_text_style))
    rows.append(current_row)

  if not rows:
    elements.append(
        Paragraph("<i>Keine Einrichtungen zum Drucken vorhanden.</i>", label_id_style)
    )
  else:
    main_table = Table(rows, colWidths=[63 * mm, 63 * mm, 63 * mm])
    main_table.setStyle(
        TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    elements.append(main_table)

  doc.build(elements)
  buffer.seek(0)
  return buffer.getvalue()


# --- EXCEL GENERATOR ---
def generate_excel_export(property_id=None):
  conn = get_db()
  wb = openpyxl.Workbook()

  header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
  header_fill = PatternFill(
      start_color="1E293B", end_color="1E293B", fill_type="solid"
  )
  data_font = Font(name="Arial", size=10)
  thin_border = Border(
      left=Side(style="thin", color="CBD5E1"),
      right=Side(style="thin", color="CBD5E1"),
      top=Side(style="thin", color="CBD5E1"),
      bottom=Side(style="thin", color="CBD5E1"),
  )
  center_align = Alignment(horizontal="center", vertical="center")
  left_align = Alignment(horizontal="left", vertical="center")

  def style_sheet(ws, title, df):
    ws.title = title
    ws.views.sheetView[0].showGridLines = True
    for col_num, col_name in enumerate(df.columns, 1):
      cell = ws.cell(row=1, column=col_num, value=str(col_name))
      cell.font = header_font
      cell.fill = header_fill
      cell.alignment = center_align
      cell.border = thin_border
      ws.row_dimensions[1].height = 24

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
      ws.row_dimensions[r_idx].height = 20
      for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = (
            center_align
            if str(val).startswith("202")
            or val in ["OK", "MANGEL", "DAUERHAFT"]
            else left_align
        )

    for col in ws.columns:
      max_len = max(len(str(cell.value or "")) for cell in col)
      col_letter = get_column_letter(col[0].column)
      ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

  q_fac = """
        SELECT p.name as Objekt, fp.name as Geschoss, f.identifier as Kennung, 
               f.category as Kategorie, f.device_type as Typ_Modell, f.location_desc as Montageort,
               f.norm_ref as Norm_Grundlage, f.last_inspection as Letzte_Pruefung, 
               f.next_inspection as Naechste_Pruefung, f.inspection_interval_months as Intervall_Monate,
               f.inspector_company as Prueffirma,
               CASE WHEN f.has_defect=1 THEN 'MANGEL' ELSE 'OK' END as Status,
               f.defect_description as Mangel_Details
        FROM fire_facilities f
        JOIN properties p ON f.property_id = p.id
        LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
    """
  params = []
  if property_id:
    q_fac += " WHERE f.property_id = ?"
    params.append(property_id)
  q_fac += " ORDER BY f.property_id, f.floor_plan_id, f.id"
  df_fac = pd.read_sql_query(q_fac, conn, params=params)

  ws1 = wb.active
  style_sheet(ws1, "Anlagenkataster", df_fac)

  q_ev = """
        SELECT p.name as Objekt, j.event_date as Datum, j.event_time as Uhrzeit, 
               j.event_type as Kategorie, j.sub_type as Klassifizierung, j.title as Betreff, 
               j.details as Hergang_Massnahmen, j.detector_group as Meldergruppe, 
               CASE WHEN j.fire_brigade_deployed=1 THEN 'JA' ELSE 'NEIN' END as FW_Einsatz,
               j.participants_count as Teilnehmer, j.duration_minutes as Dauer_Minuten, 
               j.instructor_name as Verantwortlich
        FROM journal_events j
        JOIN properties p ON j.property_id = p.id
    """
  if property_id:
    q_ev += " WHERE j.property_id = ?"
  q_ev += " ORDER BY j.event_date DESC"
  df_ev = pd.read_sql_query(q_ev, conn, params=params)
  ws2 = wb.create_sheet(title="Ereignisjournal")
  style_sheet(ws2, "Ereignisjournal", df_ev)

  q_def = """
        SELECT p.name as Objekt, fp.name as Geschoss, f.identifier as Geraet, f.category as Kategorie,
               f.location_desc as Standort, f.defect_description as Mangelbeschreibung, 
               f.defect_severity as Dringlichkeit, f.defect_due_date as Behebungsfrist
        FROM fire_facilities f
        JOIN properties p ON f.property_id = p.id
        LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
        WHERE f.has_defect = 1
    """
  if property_id:
    q_def += " AND f.property_id = ?"
  q_def += " ORDER BY f.defect_due_date ASC"
  df_def = pd.read_sql_query(q_def, conn, params=params)
  ws3 = wb.create_sheet(title="Offene_Maengel")
  style_sheet(ws3, "Offene_Maengel", df_def)

  conn.close()
  out = io.BytesIO()
  wb.save(out)
  out.seek(0)
  return out.getvalue()


# --- ZIP BACKUP & RESTORE GENERATOR ---
def create_full_backup_zip():
  zip_buffer = io.BytesIO()
  with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
    if os.path.exists(DB_FILE):
      zip_file.write(DB_FILE, arcname=DB_FILE)

    if os.path.exists(UPLOAD_DIR):
      for root, _, files in os.walk(UPLOAD_DIR):
        for file in files:
          file_path = os.path.join(root, file)
          arcname = os.path.relpath(file_path, start=".")
          zip_file.write(file_path, arcname=arcname)

  zip_buffer.seek(0)
  return zip_buffer.getvalue()


def restore_backup_from_zip(zip_bytes):
  with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zip_file:
    zip_file.extractall(".")
  init_db()


# --- PDF Generator für Handwerker-Mängelauftrag ---
def generate_handwerker_pdf(property_id, selected_defect_ids=None):
  conn = get_db()
  prop = conn.execute(
      "SELECT * FROM properties WHERE id = ?", (property_id,)
  ).fetchone()

  query = """
        SELECT f.*, fp.name as plan_name, fp.image_path as plan_image_path
        FROM fire_facilities f
        LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
        WHERE f.property_id = ? AND f.has_defect = 1
    """
  params = [property_id]
  if selected_defect_ids:
    placeholders = ",".join("?" for _ in selected_defect_ids)
    query += f" AND f.id IN ({placeholders})"
    params.extend(selected_defect_ids)

  query += " ORDER BY f.floor_plan_id, f.id ASC"
  defects = conn.execute(query, params).fetchall()
  conn.close()

  buffer = io.BytesIO()
  doc = SimpleDocTemplate(
      buffer,
      pagesize=A4,
      leftMargin=15 * mm,
      rightMargin=15 * mm,
      topMargin=15 * mm,
      bottomMargin=15 * mm,
  )

  styles = getSampleStyleSheet()
  title_style = ParagraphStyle(
      "HTitle",
      parent=styles["Heading1"],
      fontSize=15,
      leading=18,
      textColor=colors.HexColor("#991B1B"),
      spaceAfter=2,
  )
  sub_style = ParagraphStyle(
      "HSub",
      parent=styles["Normal"],
      fontSize=8.5,
      leading=11,
      textColor=colors.HexColor("#475569"),
  )
  sec_style = ParagraphStyle(
      "HSec",
      parent=styles["Heading2"],
      fontSize=11,
      leading=14,
      textColor=colors.HexColor("#1E293B"),
      spaceBefore=6,
      spaceAfter=3,
  )
  td_style = ParagraphStyle(
      "HTD",
      parent=styles["Normal"],
      fontSize=8,
      leading=10.5,
      textColor=colors.HexColor("#1E293B"),
  )

  elements = [
      Paragraph("BRANDSCHUTZ – MÄNGELBEHEBUNGSAUFTRAG", title_style),
      Paragraph(
          "Instandsetzungsauftrag & Mängelliste für Haustechnik und"
          " Fachunternehmen",
          sub_style,
      ),
      Spacer(1, 4 * mm),
  ]

  meta_data = [
      [
          Paragraph("<b>Liegenschaft / Objekt:</b>", td_style),
          Paragraph(f"{prop['name']}<br/>{prop['address']}", td_style),
          Paragraph("<b>Auftragsdatum:</b>", td_style),
          Paragraph(f"{date.today().strftime('%Y-%m-%d')}", td_style),
      ],
      [
          Paragraph("<b>Auftraggeber (BSB):</b>", td_style),
          Paragraph(f"{prop['fire_safety_officer']}", td_style),
          Paragraph("<b>Offene Mängel:</b>", td_style),
          Paragraph(
              f"<b><font color='#DC2626'>{len(defects)} Position(en)</font></b>",
              td_style,
          ),
      ],
  ]
  meta_table = Table(
      meta_data, colWidths=[35 * mm, 60 * mm, 35 * mm, 50 * mm]
  )
  meta_table.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FEF2F2")),
          ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#FCA5A5")),
          ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#FECACA")),
          ("TOPPADDING", (0, 0), (-1, -1), 4),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
      ])
  )
  elements.append(meta_table)
  elements.append(Spacer(1, 5 * mm))

  if not defects:
    elements.append(
        Paragraph("<i>Keine offenen Mängel für diesen Auftrag ausgewählt.</i>", td_style)
    )
  else:
    plans_dict = {}
    for d in defects:
      pid = d["floor_plan_id"]
      if pid:
        if pid not in plans_dict:
          plans_dict[pid] = {
              "name": d["plan_name"],
              "image_path": d["plan_image_path"],
              "defects": [],
          }
        plans_dict[pid]["defects"].append(d)

    for pid, pdata in plans_dict.items():
      plan_elem = []
      plan_elem.append(
          Paragraph(f"<b>Geschossplan: {pdata['name']}</b>", sec_style)
      )
      plan_elem.append(Spacer(1, 2 * mm))

      img_path = pdata["image_path"]
      if img_path and os.path.exists(img_path):
        base_img = Image.open(img_path).convert("RGBA")
        bw, bh = base_img.size
        target_pw = 1000
        target_ph = int((bh / bw) * target_pw)
        p_img = base_img.resize((target_pw, target_ph))

        draw = ImageDraw.Draw(p_img)
        for f in pdata["defects"]:
          if f["pin_x"] and f["pin_y"]:
            px = int((f["pin_x"] / 100.0) * target_pw)
            py = int((f["pin_y"] / 100.0) * target_ph)
            draw_color_dot(
                draw,
                px,
                py,
                fill_color="#DC2626",
                outline_color="white",
                label=f["identifier"][:2],
                radius=16,
            )

        img_byte_arr = io.BytesIO()
        p_img.convert("RGB").save(img_byte_arr, format="JPEG", quality=85)
        img_byte_arr.seek(0)

        pdf_img_w = 180 * mm
        pdf_img_h = (target_ph / target_pw) * pdf_img_w
        if pdf_img_h > 85 * mm:
          pdf_img_h = 85 * mm
          pdf_img_w = (target_pw / target_ph) * pdf_img_h

        rl_img = RLImage(img_byte_arr, width=pdf_img_w, height=pdf_img_h)
        plan_elem.append(rl_img)
        plan_elem.append(Spacer(1, 3 * mm))

      elements.append(KeepTogether(plan_elem))

      for f in pdata["defects"]:
        card_content = []
        photo_elem = Paragraph("<i>Kein Foto</i>", td_style)
        if f["defect_photo_path"] and os.path.exists(f["defect_photo_path"]):
          try:
            photo_elem = RLImage(
                f["defect_photo_path"], width=40 * mm, height=30 * mm
            )
          except Exception:
            photo_elem = Paragraph("<i>Foto nicht lesbar</i>", td_style)

        sev_col = "#DC2626" if f["defect_severity"] == "KRITISCH" else "#D97706"
        desc_text = f"""
                <b>Gerät / Kennung:</b> {f['identifier']} ({f['category']} – {f['device_type'] or ''})<br/>
                <b>Montageort:</b> {f['location_desc']}<br/>
                <b>Mangelbeschreibung:</b> <b>{f['defect_description'] or 'Keine Angabe'}</b><br/>
                <b>Dringlichkeit:</b> <font color='{sev_col}'><b>{f['defect_severity'] or 'MITTEL'}</b></font> | <b>Frist bis:</b> {f['defect_due_date'] or '-'}<br/>
                """

        card_table_data = [[
            Paragraph(desc_text, td_style),
            photo_elem,
            Paragraph(
                "<b>Rückmeldung Handwerker:</b><br/>[ ] Behoben"
                " am:_________<br/>Name:________________<br/>Unterschrift:__________",
                td_style,
            ),
        ]]
        card_table = Table(
            card_table_data, colWidths=[90 * mm, 45 * mm, 45 * mm]
        )
        card_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])
        )
        card_content.append(card_table)
        card_content.append(Spacer(1, 3 * mm))
        elements.append(KeepTogether(card_content))

  elements.append(Spacer(1, 8 * mm))
  sign_data = [[
      Paragraph(
          "____________________________________________<br/>Datum, Unterschrift"
          " Auftragsvergabe (BSB)",
          td_style,
      ),
      Paragraph(
          "____________________________________________<br/>Datum, Unterschrift"
          " Fertigstellung & Abnahme",
          td_style,
      ),
  ]]
  s_table = Table(sign_data, colWidths=[90 * mm, 90 * mm])
  s_table.setStyle(
      TableStyle([
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("LEFTPADDING", (0, 0), (-1, -1), 0),
      ])
  )
  elements.append(s_table)

  doc.build(elements)
  buffer.seek(0)
  return buffer.getvalue()


# --- PDF Generator für Begehungsprotokoll ---
def generate_combined_pdf(run_id):
  conn = get_db()
  run = conn.execute(
      """
        SELECT r.*, p.name as prop_name, p.address as prop_address, p.fire_safety_officer 
        FROM inspection_runs r
        JOIN properties p ON r.property_id = p.id
        WHERE r.id = ?
    """,
      (run_id,),
  ).fetchone()

  items = conn.execute(
      """
        SELECT t.facility_category, t.trvb_ref, t.check_item, ir.result_status, ir.remark
        FROM inspection_results ir
        JOIN checklist_templates t ON ir.checklist_template_id = t.id
        WHERE ir.inspection_run_id = ?
        ORDER BY t.facility_category, t.id
    """,
      (run_id,),
  ).fetchall()

  facilities = conn.execute(
      """
        SELECT f.*, fp.name as plan_name, fp.image_path as plan_image_path
        FROM fire_facilities f
        LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
        WHERE f.property_id = ?
        ORDER BY f.floor_plan_id, f.id ASC
    """,
      (run["property_id"],),
  ).fetchall()
  conn.close()

  buffer = io.BytesIO()
  doc = SimpleDocTemplate(
      buffer,
      pagesize=A4,
      leftMargin=15 * mm,
      rightMargin=15 * mm,
      topMargin=15 * mm,
      bottomMargin=15 * mm,
  )

  styles = getSampleStyleSheet()
  title_style = ParagraphStyle(
      "TitleStyle",
      parent=styles["Heading1"],
      fontSize=15,
      leading=18,
      textColor=colors.HexColor("#0F172A"),
      spaceAfter=2,
  )
  section_style = ParagraphStyle(
      "SecTitle",
      parent=styles["Heading2"],
      fontSize=11,
      leading=14,
      textColor=colors.HexColor("#1E293B"),
      spaceBefore=6,
      spaceAfter=4,
  )
  sub_style = ParagraphStyle(
      "SubTitleStyle",
      parent=styles["Normal"],
      fontSize=8.5,
      leading=11,
      textColor=colors.HexColor("#475569"),
  )
  th_style = ParagraphStyle(
      "TH",
      parent=styles["Normal"],
      fontSize=8,
      leading=10,
      fontName="Helvetica-Bold",
      textColor=colors.white,
  )
  td_style = ParagraphStyle(
      "TD",
      parent=styles["Normal"],
      fontSize=7.5,
      leading=9.5,
      textColor=colors.HexColor("#1E293B"),
  )

  elements = [
      Paragraph("BRANDSCHUTZBUCH – EIGENKONTROLLPROTOKOLL", title_style),
      Paragraph(
          "Kombinierte Begehung, Anlagenprüfung & Mängelprotokoll gem. TRVB 117"
          " O",
          sub_style,
      ),
      Spacer(1, 4 * mm),
  ]

  meta_data = [
      [
          Paragraph("<b>Objekt:</b>", td_style),
          Paragraph(
              f"{run['prop_name']}<br/>{run['prop_address']}", td_style
          ),
          Paragraph("<b>Prüfdatum:</b>", td_style),
          Paragraph(f"{run['inspection_date']}", td_style),
      ],
      [
          Paragraph("<b>Prüfer / BSB:</b>", td_style),
          Paragraph(f"{run['inspector_name']}", td_style),
          Paragraph("<b>Intervall:</b>", td_style),
          Paragraph(f"<b>{run['interval_scope']}</b>", td_style),
      ],
  ]
  meta_table = Table(
      meta_data, colWidths=[25 * mm, 65 * mm, 25 * mm, 65 * mm]
  )
  meta_table.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
          ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
          ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
          ("TOPPADDING", (0, 0), (-1, -1), 3),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
      ])
  )
  elements.append(meta_table)
  elements.append(Spacer(1, 4 * mm))

  elements.append(
      Paragraph("1. TRVB 117 O Prüfstatus nach Kategorien", section_style)
  )
  t_data = [[
      Paragraph("Bereich", th_style),
      Paragraph("Prüfhandlung / Anforderung", th_style),
      Paragraph("Status", th_style),
      Paragraph("Bemerkung / Feststellungen", th_style),
  ]]
  for it in items:
    status = it["result_status"]
    color = (
        "#DC2626"
        if status == "MANGEL"
        else ("#16A34A" if status == "OK" else "#64748B")
    )
    t_data.append([
        Paragraph(
            f"<b>{it['facility_category']}</b><br/><font"
            f" color='#64748B'>{it['trvb_ref']}</font>",
            td_style,
        ),
        Paragraph(it["check_item"], td_style),
        Paragraph(f"<b><font color='{color}'>{status}</font></b>", td_style),
        Paragraph(it["remark"] if it["remark"] else "-", td_style),
    ])
  t_res = Table(t_data, colWidths=[35 * mm, 75 * mm, 20 * mm, 50 * mm])
  t_res.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
          ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
          ("TOPPADDING", (0, 0), (-1, -1), 3),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
      ])
  )
  elements.append(t_res)
  elements.append(Spacer(1, 5 * mm))

  elements.append(
      Paragraph(
          "2. Plan-Übersicht der Brandschutzeinrichtungen & Mängel",
          section_style,
      )
  )

  plans_dict = {}
  for f in facilities:
    pid = f["floor_plan_id"]
    if pid:
      if pid not in plans_dict:
        plans_dict[pid] = {
            "name": f["plan_name"],
            "image_path": f["plan_image_path"],
            "facilities": [],
        }
      plans_dict[pid]["facilities"].append(f)

  today = date.today()
  warn_threshold = today + timedelta(days=60)

  for pid, pdata in plans_dict.items():
    plan_elements = []
    plan_elements.append(
        Paragraph(f"<b>Geschossplan: {pdata['name']}</b>", td_style)
    )
    plan_elements.append(Spacer(1, 2 * mm))

    img_path = pdata["image_path"]
    if img_path and os.path.exists(img_path):
      base_img = Image.open(img_path).convert("RGBA")
      bw, bh = base_img.size
      target_pw = 1000
      target_ph = int((bh / bw) * target_pw)
      p_img = base_img.resize((target_pw, target_ph))

      draw = ImageDraw.Draw(p_img)
      for f in pdata["facilities"]:
        if f["pin_x"] and f["pin_y"]:
          px = int((f["pin_x"] / 100.0) * target_pw)
          py = int((f["pin_y"] / 100.0) * target_ph)

          next_dt = None
          if (
              f["next_inspection"]
              and f["next_inspection"] not in ["DAUERHAFT", "None", ""]
          ):
            try:
              next_dt = datetime.strptime(
                  f["next_inspection"], "%Y-%m-%d"
              ).date()
            except ValueError:
              pass

          if f["has_defect"] or (next_dt and next_dt < today):
            p_color = "#DC2626"
          elif next_dt and next_dt <= warn_threshold:
            p_color = "#D97706"
          else:
            p_color = "#16A34A"

          draw_color_dot(
              draw,
              px,
              py,
              fill_color=p_color,
              outline_color="white",
              label=f["identifier"][:2],
              radius=15,
          )

      img_byte_arr = io.BytesIO()
      p_img.convert("RGB").save(img_byte_arr, format="JPEG", quality=85)
      img_byte_arr.seek(0)

      pdf_img_w = 180 * mm
      pdf_img_h = (target_ph / target_pw) * pdf_img_w
      if pdf_img_h > 95 * mm:
        pdf_img_h = 95 * mm
        pdf_img_w = (target_pw / target_ph) * pdf_img_h

      rl_img = RLImage(img_byte_arr, width=pdf_img_w, height=pdf_img_h)
      plan_elements.append(rl_img)
      plan_elements.append(Spacer(1, 3 * mm))

    f_data = [[
        Paragraph("Kennung", th_style),
        Paragraph("Kategorie / Typ", th_style),
        Paragraph("Standort", th_style),
        Paragraph("Prüffrist", th_style),
        Paragraph("Status / Mangel", th_style),
    ]]
    for f in pdata["facilities"]:
      next_dt = None
      if (
          f["next_inspection"]
          and f["next_inspection"] not in ["DAUERHAFT", "None", ""]
      ):
        try:
          next_dt = datetime.strptime(
              f["next_inspection"], "%Y-%m-%d"
          ).date()
        except ValueError:
          pass

      if f["has_defect"]:
        st_txt = f"<b><font color='#DC2626'>Mangel: {f['defect_description'] or 'vorhanden'}</font></b>"
      elif next_dt and next_dt < today:
        st_txt = "<b><font color='#DC2626'>Überfällig</font></b>"
      elif next_dt and next_dt <= warn_threshold:
        st_txt = "<b><font color='#D97706'>Fällig <60d</font></b>"
      else:
        st_txt = "<b><font color='#16A34A'>Gültig / i.O.</font></b>"

      frist_txt = (
          f"{f['next_inspection']}"
          if f["next_inspection"] not in ["DAUERHAFT", "None", ""]
          else "Keine Frist"
      )

      f_data.append([
          Paragraph(f"<b>{f['identifier']}</b>", td_style),
          Paragraph(f"{f['category']}<br/>{f['device_type'] or ''}", td_style),
          Paragraph(f"{f['location_desc']}", td_style),
          Paragraph(frist_txt, td_style),
          Paragraph(st_txt, td_style),
      ])
    t_pfac = Table(
        f_data, colWidths=[20 * mm, 35 * mm, 55 * mm, 25 * mm, 45 * mm]
    )
    t_pfac.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])
    )
    plan_elements.append(t_pfac)
    plan_elements.append(Spacer(1, 5 * mm))

    elements.append(KeepTogether(plan_elements))

  elements.append(Spacer(1, 6 * mm))

  sig_bsb_elem = Paragraph("<i>Keine digitale Signatur</i>", td_style)
  if run["sig_bsb_path"] and os.path.exists(run["sig_bsb_path"]):
    sig_bsb_elem = RLImage(run["sig_bsb_path"], width=45 * mm, height=18 * mm)

  sig_mgmt_elem = Paragraph("<i>Keine digitale Signatur</i>", td_style)
  if run["sig_mgmt_path"] and os.path.exists(run["sig_mgmt_path"]):
    sig_mgmt_elem = RLImage(run["sig_mgmt_path"], width=45 * mm, height=18 * mm)

  sign_data = [
      [sig_bsb_elem, sig_mgmt_elem],
      [
          Paragraph(
              "____________________________________________<br/>Datum,"
              " Unterschrift Brandschutzbeauftragter",
              td_style,
          ),
          Paragraph(
              "____________________________________________<br/>Kenntnisnahme"
              " Betriebsleitung / Eigentümer",
              td_style,
          ),
      ],
  ]
  s_table = Table(sign_data, colWidths=[90 * mm, 90 * mm])
  s_table.setStyle(
      TableStyle([
          ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
          ("LEFTPADDING", (0, 0), (-1, -1), 0),
          ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
      ])
  )
  elements.append(s_table)

  doc.build(elements)
  buffer.seek(0)
  return buffer.getvalue()


# --- PDF Generator für TRVB 117 O Ereignisjournal ---
def generate_journal_pdf(property_id):
  conn = get_db()
  prop = conn.execute(
      "SELECT * FROM properties WHERE id = ?", (property_id,)
  ).fetchone()
  events = conn.execute(
      """
        SELECT * FROM journal_events 
        WHERE property_id = ? 
        ORDER BY event_date DESC, id DESC
    """,
      (property_id,),
  ).fetchall()
  conn.close()

  buffer = io.BytesIO()
  doc = SimpleDocTemplate(
      buffer,
      pagesize=A4,
      leftMargin=15 * mm,
      rightMargin=15 * mm,
      topMargin=15 * mm,
      bottomMargin=15 * mm,
  )

  styles = getSampleStyleSheet()
  title_style = ParagraphStyle(
      "TitleStyle",
      parent=styles["Heading1"],
      fontSize=15,
      leading=18,
      textColor=colors.HexColor("#0F172A"),
      spaceAfter=2,
  )
  sub_style = ParagraphStyle(
      "SubTitleStyle",
      parent=styles["Normal"],
      fontSize=8.5,
      leading=11,
      textColor=colors.HexColor("#475569"),
  )
  th_style = ParagraphStyle(
      "TH",
      parent=styles["Normal"],
      fontSize=8,
      leading=10,
      fontName="Helvetica-Bold",
      textColor=colors.white,
  )
  td_style = ParagraphStyle(
      "TD",
      parent=styles["Normal"],
      fontSize=7.5,
      leading=9.5,
      textColor=colors.HexColor("#1E293B"),
  )

  elements = [
      Paragraph("BRANDSCHUTZBUCH – EREIGNISJOURNAL", title_style),
      Paragraph(
          "Offizielles Protokoll über Alarme, Übungen & Schulungen gem. TRVB"
          " 117 O",
          sub_style,
      ),
      Spacer(1, 4 * mm),
  ]

  meta_data = [
      [
          Paragraph("<b>Liegenschaft / Betrieb:</b>", td_style),
          Paragraph(f"{prop['name']}<br/>{prop['address']}", td_style),
          Paragraph("<b>Druckdatum:</b>", td_style),
          Paragraph(f"{date.today().strftime('%Y-%m-%d')}", td_style),
      ],
      [
          Paragraph("<b>Verantw. BSB:</b>", td_style),
          Paragraph(f"{prop['fire_safety_officer']}", td_style),
          Paragraph("<b>Dokumentierte Ereignisse:</b>", td_style),
          Paragraph(f"<b>{len(events)} Einträge</b>", td_style),
      ],
  ]
  meta_table = Table(
      meta_data, colWidths=[35 * mm, 60 * mm, 35 * mm, 50 * mm]
  )
  meta_table.setStyle(
      TableStyle([
          ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
          ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
          ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
          ("TOPPADDING", (0, 0), (-1, -1), 3),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
      ])
  )
  elements.append(meta_table)
  elements.append(Spacer(1, 6 * mm))

  if not events:
    elements.append(
        Paragraph("<i>Bisher keine Ereignisse im Journal erfasst.</i>", td_style)
    )
  else:
    j_data = [[
        Paragraph("Datum / Zeit", th_style),
        Paragraph("Ereignisart", th_style),
        Paragraph("Bezeichnung & Details", th_style),
        Paragraph("Zusatzdaten", th_style),
        Paragraph("Verantwortlich", th_style),
    ]]

    for ev in events:
      time_txt = f"<br/>{ev['event_time']} Uhr" if ev["event_time"] else ""
      sub_txt = f"<br/><i>({ev['sub_type']})</i>" if ev["sub_type"] else ""

      extra_lines = []
      if ev["detector_group"]:
        extra_lines.append(f"Meldergruppe: {ev['detector_group']}")
      if ev["fire_brigade_deployed"]:
        extra_lines.append("<b><font color='#DC2626'>🚒 FW-Einsatz</font></b>")
      if ev["participants_count"]:
        extra_lines.append(f"Teilnehmer: {ev['participants_count']}")
      if ev["duration_minutes"]:
        extra_lines.append(f"Dauer/Zeit: {ev['duration_minutes']} min")
      extra_txt = "<br/>".join(extra_lines) if extra_lines else "-"

      j_data.append([
          Paragraph(f"<b>{ev['event_date']}</b>{time_txt}", td_style),
          Paragraph(f"<b>{ev['event_type']}</b>{sub_txt}", td_style),
          Paragraph(
              f"<b>{ev['title']}</b><br/>{ev['details'] or ''}", td_style
          ),
          Paragraph(extra_txt, td_style),
          Paragraph(ev["instructor_name"] or "-", td_style),
      ])

    j_table = Table(
        j_data, colWidths=[24 * mm, 30 * mm, 66 * mm, 32 * mm, 28 * mm]
    )
    j_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [colors.white, colors.HexColor("#F8FAFC")],
            ),
        ])
    )
    elements.append(j_table)

  elements.append(Spacer(1, 12 * mm))
  sign_data = [[
      Paragraph(
          "____________________________________________<br/>Datum, Unterschrift"
          " Brandschutzbeauftragter",
          td_style,
      ),
      Paragraph(
          "____________________________________________<br/>Kenntnisnahme"
          " Betriebsleitung / Eigentümer",
          td_style,
      ),
  ]]
  s_table = Table(sign_data, colWidths=[90 * mm, 90 * mm])
  s_table.setStyle(
      TableStyle([
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("LEFTPADDING", (0, 0), (-1, -1), 0),
      ])
  )
  elements.append(s_table)

  doc.build(elements)
  buffer.seek(0)
  return buffer.getvalue()


# --- Streamlit Navigation ---
st.set_page_config(
    page_title="BSB Manager TRVB 117 O", layout="wide", page_icon="🧯"
)
st.title("🧯 BSB Manager – Brandschutzbuch & Begehung")

menu = st.sidebar.radio(
    "Navigation",
    [
        "🎨 Objektpläne & Kataster bearbeiten",
        "Aktive Begehung & Planprüfung",
        "📷 Live-Kamera QR-Scanner",
        "Anlagenkataster & Fristen",
        "🏷️ QR-Code Etikettendruck",
        "🛠️ Handwerker- & Mängelauftrag",
        "🚨 Ereignisjournal (TRVB 117 O)",
        "📊 BSB-Jahresbericht (TRVB 117 O)",
        "💾 Daten-Export & Datensicherung",
        "Brandschutzbuch-Historie",
        "Pläne & Objekte verwalten",
    ],
)
conn = get_db()


# ----------------------------------------------------
# 1. OBJEKTPLÄNE & KATASTER BEARBEITEN
# ----------------------------------------------------
if menu == "🎨 Objektpläne & Kataster bearbeiten":
  st.subheader("🎨 Objektpläne bearbeiten & Einrichtungen platzieren")
  st.caption(
      "Platziere Brandschutzeinrichtungen per Klick auf den Plan oder bearbeite"
      " bestehende Punkte."
  )

  properties = conn.execute("SELECT * FROM properties").fetchall()
  if not properties:
    st.warning("Bitte zuerst ein Objekt unter 'Pläne & Objekte' anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox(
        "Objekt auswählen", list(prop_dict.keys()), key="edit_prop"
    )
    selected_prop_id = prop_dict[selected_prop_label]

    plans = conn.execute(
        "SELECT * FROM floor_plans WHERE property_id = ?", (selected_prop_id,)
    ).fetchall()

    if not plans:
      st.info(
          "Für dieses Objekt ist noch kein Plan hinterlegt. Lade Pläne unter"
          " 'Pläne & Objekte verwalten' hoch."
      )
    else:
      plan_dict = {pl["name"]: pl["id"] for pl in plans}
      selected_plan_name = st.selectbox(
          "Geschossplan zum Bearbeiten auswählen",
          list(plan_dict.keys()),
          key="edit_plan_select",
      )
      selected_plan_id = plan_dict[selected_plan_name]

      curr_plan = conn.execute(
          "SELECT * FROM floor_plans WHERE id = ?", (selected_plan_id,)
      ).fetchone()
      plan_img_path = curr_plan["image_path"]

      if os.path.exists(plan_img_path):
        base_img = Image.open(plan_img_path).convert("RGBA")
        w, h = base_img.size
        tw = 950
        th = int((h / w) * tw)
        r_img = base_img.resize((tw, th))

        facs_on_plan = conn.execute(
            "SELECT * FROM fire_facilities WHERE floor_plan_id = ?",
            (selected_plan_id,),
        ).fetchall()

        d_img = r_img.copy()
        draw = ImageDraw.Draw(d_img)

        for fac in facs_on_plan:
          if fac["pin_x"] and fac["pin_y"]:
            fx = int((fac["pin_x"] / 100.0) * tw)
            fy = int((fac["pin_y"] / 100.0) * th)
            draw_color_dot(
                draw,
                fx,
                fy,
                fill_color="#2563EB",
                outline_color="white",
                label=fac["category"][:1],
                radius=13,
            )
            draw.text(
                (fx - 12, fy + 14),
                fac["identifier"],
                fill="black",
                stroke_fill="white",
                stroke_width=2,
            )

        col_plan_view, col_plan_form = st.columns([6.5, 3.5])

        with col_plan_view:
          st.caption(
              f"📍 Klicke auf den Plan: **Freie Fläche = Neuer Farbpunkt** |"
              f" **Bestehender Punkt = Bearbeiten/Löschen**"
          )
          edit_coords = streamlit_image_coordinates(
              d_img, key=f"edit_coords_{selected_plan_id}"
          )

        with col_plan_form:
          st.markdown("#### Einrichtung konfigurieren")
          if edit_coords:
            px_val = round((edit_coords["x"] / tw) * 100, 2)
            py_val = round((edit_coords["y"] / th) * 100, 2)

            clicked_fac = None
            for fac in facs_on_plan:
              if fac["pin_x"] and fac["pin_y"]:
                if (
                    abs(fac["pin_x"] - px_val) < 4.0
                    and abs(fac["pin_y"] - py_val) < 4.0
                ):
                  clicked_fac = fac
                  break

            if clicked_fac:
              st.success(
                  f"📍 Bearbeite: **{clicked_fac['identifier']}**"
                  f" ({clicked_fac['category']})"
              )

              with st.form("edit_existing_fac"):
                e_ident = st.text_input(
                    "Kennung / Gerätenummer", value=clicked_fac["identifier"]
                )
                e_cat = st.selectbox(
                    "Kategorie",
                    [
                        "LOESCHER",
                        "TUER",
                        "HYDRANT",
                        "BMA",
                        "RWA",
                        "NOTLICHT",
                        "FLUECHTWEG",
                        "SONSTIGES",
                    ],
                    index=[
                        "LOESCHER",
                        "TUER",
                        "HYDRANT",
                        "BMA",
                        "RWA",
                        "NOTLICHT",
                        "FLUECHTWEG",
                        "SONSTIGES",
                    ].index(clicked_fac["category"]),
                )
                e_type = st.text_input(
                    "Typ / Modell", value=clicked_fac["device_type"] or ""
                )
                e_loc = st.text_input(
                    "Standortbeschreibung", value=clicked_fac["location_desc"]
                )

                has_no_inspection = (
                    clicked_fac["inspection_interval_months"] == 0
                    or clicked_fac["next_inspection"]
                    in ["DAUERHAFT", "None", ""]
                )
                e_no_insp = st.checkbox(
                    "Keine wiederkehrende Prüfung (dauerhaft gültig)",
                    value=has_no_inspection,
                )

                if not e_no_insp:
                  try:
                    last_val = datetime.strptime(
                        clicked_fac["last_inspection"], "%Y-%m-%d"
                    ).date()
                  except Exception:
                    last_val = date.today()

                  e_last = st.date_input("Letzte Prüfung", value=last_val)
                  e_int = st.number_input(
                      "Prüfintervall (Monate)",
                      value=clicked_fac["inspection_interval_months"]
                      if clicked_fac["inspection_interval_months"] > 0
                      else 24,
                  )
                else:
                  e_last = None
                  e_int = 0

                col_b1, col_b2 = st.columns(2)
                if col_b1.form_submit_button(
                    "💾 Änderungen speichern", type="primary"
                ):
                  if e_no_insp:
                    final_last_str = "DAUERHAFT"
                    final_next_str = "DAUERHAFT"
                  else:
                    final_last_str = e_last.strftime("%Y-%m-%d")
                    e_next = (
                        e_last.replace(year=e_last.year + (e_int // 12))
                        if e_int % 12 == 0
                        else e_last + timedelta(days=e_int * 30)
                    )
                    final_next_str = e_next.strftime("%Y-%m-%d")

                  conn.execute(
                      """
                                            UPDATE fire_facilities 
                                            SET identifier = ?, category = ?, device_type = ?, location_desc = ?,
                                                last_inspection = ?, next_inspection = ?, inspection_interval_months = ?
                                            WHERE id = ?
                                        """,
                      (
                          e_ident,
                          e_cat,
                          e_type,
                          e_loc,
                          final_last_str,
                          final_next_str,
                          e_int,
                          clicked_fac["id"],
                      ),
                  )
                  conn.commit()
                  st.success("Aktualisiert!")
                  st.rerun()

              st.divider()
              del_confirm = st.checkbox(
                  "Löschen bestätigen", key=f"chk_del_fac_{clicked_fac['id']}"
              )
              if st.button(
                  "🗑️ Einrichtung endgültig löschen",
                  type="secondary",
                  disabled=not del_confirm,
              ):
                conn.execute(
                    "DELETE FROM fire_facilities WHERE id = ?",
                    (clicked_fac["id"],),
                )
                conn.commit()
                st.warning(
                    f"Einrichtung '{clicked_fac['identifier']}' wurde gelöscht!"
                )
                st.rerun()

            else:
              st.info(
                  f"📍 Neuer Farbpunkt an Position **X: {px_val}%, Y:"
                  f" {py_val}%**"
              )
              with st.form("new_fac_on_plan", clear_on_submit=True):
                cat = st.selectbox(
                    "Kategorie*",
                    [
                        "LOESCHER",
                        "TUER",
                        "HYDRANT",
                        "BMA",
                        "RWA",
                        "NOTLICHT",
                        "FLUECHTWEG",
                        "SONSTIGES",
                    ],
                )
                ident = st.text_input(
                    "Gerätenummer / Kennung*",
                    placeholder="z.B. FL-01 oder T30-EG-01",
                )
                dev_t = st.text_input(
                    "Typ / Modell / Löschmittel",
                    placeholder="z.B. 6kg Schaum / T30-RS",
                )
                loc = st.text_input(
                    "Standortbeschreibung*",
                    placeholder="z.B. Flur West neben Ausgang",
                )

                no_insp = st.checkbox(
                    "Keine wiederkehrende Prüfung (dauerhaft gültig / nur"
                    " Sichtprüfung)"
                )
                last_i = st.date_input("Letzte Fachprüfung", value=date.today())
                int_m = st.number_input(
                    "Prüfintervall (Monate)",
                    value=24 if cat == "LOESCHER" else 12,
                )

                if st.form_submit_button(
                    "➕ Einrichtung hier platzieren", type="primary"
                ):
                  if ident and loc:
                    if no_insp:
                      final_last_str = "DAUERHAFT"
                      final_next_str = "DAUERHAFT"
                      final_int = 0
                    else:
                      final_last_str = last_i.strftime("%Y-%m-%d")
                      final_int = int_m
                      calc_next = (
                          last_i.replace(year=last_i.year + (int_m // 12))
                          if int_m % 12 == 0
                          else last_i + timedelta(days=int_m * 30)
                      )
                      final_next_str = calc_next.strftime("%Y-%m-%d")

                    conn.execute(
                        """
                                                INSERT INTO fire_facilities (
                                                    property_id, floor_plan_id, pin_x, pin_y, category, identifier, 
                                                    device_type, location_desc, last_inspection, next_inspection, 
                                                    inspection_interval_months
                                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                            """,
                        (
                            selected_prop_id,
                            selected_plan_id,
                            px_val,
                            py_val,
                            cat,
                            ident,
                            dev_t,
                            loc,
                            final_last_str,
                            final_next_str,
                            final_int,
                        ),
                    )
                    conn.commit()
                    st.success(f"Einrichtung '{ident}' erfolgreich platziert!")
                    st.rerun()
                  else:
                    st.error("Bitte Kennung und Standort ausfüllen.")
          else:
            st.info("Klicke links auf den Plan, um ein Gerät zu platzieren.")

        if facs_on_plan:
          st.divider()
          st.markdown(
              f"#### Alle Einrichtungen auf Plan **{selected_plan_name}**"
              f" ({len(facs_on_plan)}):"
          )
          for f in facs_on_plan:
            frist_str = (
                f"`{f['next_inspection']}`"
                if f["next_inspection"] not in ["DAUERHAFT", "None", ""]
                else "*Keine Prüffrist*"
            )
            st.write(
                f"🔵 **{f['identifier']}** | `{f['category']}` |"
                f" {f['device_type'] or '-'} | Ort: {f['location_desc']} |"
                f" Nächste Prüfung: {frist_str}"
            )


# ----------------------------------------------------
# 2. AKTIVE BEGEHUNG & PLANPRÜFUNG
# ----------------------------------------------------
elif menu == "Aktive Begehung & Planprüfung":
  properties = conn.execute("SELECT * FROM properties").fetchall()

  if not properties:
    st.warning("Bitte zuerst ein Objekt unter 'Pläne & Objekte' anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox(
        "Objekt auswählen", list(prop_dict.keys()), key="main_prop"
    )
    selected_prop_id = prop_dict[selected_prop_label]

    col_s1, col_s2, col_s3 = st.columns(3)
    inspector = col_s1.text_input("Prüfer / BSB", value="Johannes Probst")
    inspect_date = col_s2.date_input("Prüfdatum", value=date.today())
    interval = col_s3.selectbox(
        "Intervall",
        ["MONATLICH", "WÖCHENTLICH", "QUARTAL", "HALBJÄHRLICH", "JÄHRLICH"],
    )

    if "active_run_id" not in st.session_state:
      st.session_state["active_run_id"] = None

    if st.session_state["active_run_id"] is None:
      if st.button("🚀 Neue Begehung für dieses Objekt starten", type="primary"):
        cur = conn.cursor()
        cur.execute(
            """
                    INSERT INTO inspection_runs (property_id, inspector_name, inspection_date, interval_scope, status)
                    VALUES (?, ?, ?, ?, 'IN_BEGEHUNG')
                """,
            (selected_prop_id, inspector, inspect_date, interval),
        )
        st.session_state["active_run_id"] = cur.lastrowid
        conn.commit()
        st.rerun()

    if st.session_state["active_run_id"] is not None:
      run_id = st.session_state["active_run_id"]
      st.success(
          f"🔴 **Begehung aktiv (Prüflauf #{run_id})** für {selected_prop_label}"
      )

      tab_plan, tab_check, tab_sign = st.tabs([
          "📍 1. Plan-Begehung (Punkte prüfen)",
          "📋 2. TRVB-Prüfpunkte",
          "✍️ 3. Digitale Unterschrift & Abschluss",
      ])

      with tab_plan:
        plans = conn.execute(
            "SELECT * FROM floor_plans WHERE property_id = ?",
            (selected_prop_id,),
        ).fetchall()

        if not plans:
          st.info("Keine Pläne für dieses Objekt hinterlegt.")
        else:
          plan_dict = {pl["name"]: pl["id"] for pl in plans}
          selected_plan_name = st.selectbox(
              "Geschoss / Plan auswählen (z.B. KG, EG, 1.OG)",
              list(plan_dict.keys()),
          )
          selected_plan_id = plan_dict[selected_plan_name]

          curr_plan = conn.execute(
              "SELECT * FROM floor_plans WHERE id = ?", (selected_plan_id,)
          ).fetchone()
          plan_img_path = curr_plan["image_path"]

          if os.path.exists(plan_img_path):
            base_img = Image.open(plan_img_path).convert("RGBA")
            w, h = base_img.size
            tw = 950
            th = int((h / w) * tw)
            r_img = base_img.resize((tw, th))

            facs_on_plan = conn.execute(
                "SELECT * FROM fire_facilities WHERE floor_plan_id = ?",
                (selected_plan_id,),
            ).fetchall()

            d_img = r_img.copy()
            draw = ImageDraw.Draw(d_img)

            today = date.today()
            warn_threshold = today + timedelta(days=60)

            for fac in facs_on_plan:
              if fac["pin_x"] and fac["pin_y"]:
                fx = int((fac["pin_x"] / 100.0) * tw)
                fy = int((fac["pin_y"] / 100.0) * th)

                next_dt = None
                if (
                    fac["next_inspection"]
                    and fac["next_inspection"] not in ["DAUERHAFT", "None", ""]
                ):
                  try:
                    next_dt = datetime.strptime(
                        fac["next_inspection"], "%Y-%m-%d"
                    ).date()
                  except ValueError:
                    pass

                if fac["has_defect"] or (next_dt and next_dt < today):
                  dot_col = "#DC2626"
                elif next_dt and next_dt <= warn_threshold:
                  dot_col = "#D97706"
                else:
                  dot_col = "#16A34A"

                draw_color_dot(
                    draw,
                    fx,
                    fy,
                    fill_color=dot_col,
                    outline_color="white",
                    label=fac["category"][:1],
                    radius=13,
                )
                draw.text(
                    (fx - 12, fy + 14),
                    fac["identifier"],
                    fill="black",
                    stroke_fill="white",
                    stroke_width=2,
                )

            col_p, col_f = st.columns([6.5, 3.5])
            with col_p:
              st.caption(
                  f"📍 Plan: **{selected_plan_name}** (🟢 Gültig/Dauerhaft | 🟡"
                  " <60 Tage | 🔴 Mangel/Überfällig)"
              )
              coords = streamlit_image_coordinates(
                  d_img, key=f"pin_run_{run_id}_{selected_plan_id}"
              )

            with col_f:
              st.markdown("#### Punkt prüfen / Mangel aufnehmen")
              if coords:
                px_val = round((coords["x"] / tw) * 100, 2)
                py_val = round((coords["y"] / th) * 100, 2)

                clicked_fac = None
                for fac in facs_on_plan:
                  if fac["pin_x"] and fac["pin_y"]:
                    if (
                        abs(fac["pin_x"] - px_val) < 4.0
                        and abs(fac["pin_y"] - py_val) < 4.0
                    ):
                      clicked_fac = fac
                      break

                if clicked_fac:
                  st.success(
                      f"📍 Angeklickt: **{clicked_fac['identifier']}**"
                      f" ({clicked_fac['category']})"
                  )
                  st.write(
                      f"Standort: {clicked_fac['location_desc']} | Typ:"
                      f" {clicked_fac['device_type'] or '-'}"
                  )
                  frist_disp = (
                      f"`{clicked_fac['next_inspection']}`"
                      if clicked_fac["next_inspection"]
                      not in ["DAUERHAFT", "None", ""]
                      else "*Keine wiederkehrende Frist*"
                  )
                  st.write(f"Nächste Prüfung: {frist_disp}")

                  act = st.radio(
                      "Aktion für dieses Gerät:",
                      [
                          "✅ Alles i.O. (Prüfung verlängern)",
                          "⚠️ Mangel erfassen",
                          "🔄 Mangel aufheben",
                      ],
                  )

                  if act == "✅ Alles i.O. (Prüfung verlängern)":
                    if clicked_fac["inspection_interval_months"] > 0:
                      if st.button("Plakette erneuern / Frist verlängern"):
                        new_last = date.today()
                        months = clicked_fac["inspection_interval_months"]
                        new_next = (
                            new_last.replace(
                                year=new_last.year + (months // 12)
                            )
                            if months % 12 == 0
                            else new_last + timedelta(days=months * 30)
                        )
                        conn.execute(
                            """
                                                    UPDATE fire_facilities 
                                                    SET last_inspection = ?, next_inspection = ?, has_defect = 0 
                                                    WHERE id = ?
                                                """,
                            (
                                new_last.strftime("%Y-%m-%d"),
                                new_next.strftime("%Y-%m-%d"),
                                clicked_fac["id"],
                            ),
                        )
                        conn.commit()
                        st.success(f"Gültig bis {new_next}!")
                        st.rerun()
                    else:
                      if st.button("Sichtprüfung i.O. bestätigen"):
                        conn.execute(
                            """
                                                    UPDATE fire_facilities 
                                                    SET has_defect = 0 
                                                    WHERE id = ?
                                                """,
                            (clicked_fac["id"],),
                        )
                        conn.commit()
                        st.success("Gerät in Ordnung!")
                        st.rerun()

                  elif act == "⚠️ Mangel erfassen":
                    with st.form("defect_existing_form"):
                      m_desc = st.text_area(
                          "Mangelbeschreibung",
                          placeholder="z.B. Fluchttür klemmt, Löscher verstellt",
                      )
                      m_sev = st.selectbox(
                          "Dringlichkeit", ["MITTEL", "KRITISCH", "GERING"]
                      )
                      m_due = st.date_input("Frist", value=date.today())
                      m_foto = st.file_uploader(
                          "Foto", type=["jpg", "png", "jpeg"]
                      )

                      if st.form_submit_button("Mangel speichern"):
                        f_path = clicked_fac["defect_photo_path"]
                        if m_foto:
                          fn = f"mangel_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{m_foto.name}"
                          f_path = os.path.join(UPLOAD_DIR, fn)
                          with open(f_path, "wb") as f:
                            f.write(m_foto.getbuffer())

                        conn.execute(
                            """
                                                    UPDATE fire_facilities 
                                                    SET has_defect = 1, defect_description = ?, defect_severity = ?, 
                                                        defect_due_date = ?, defect_photo_path = ?
                                                    WHERE id = ?
                                                """,
                            (
                                m_desc,
                                m_sev,
                                m_due,
                                f_path,
                                clicked_fac["id"],
                            ),
                        )
                        conn.commit()
                        st.success("Mangel am Gerät hinterlegt!")
                        st.rerun()

                  elif act == "🔄 Mangel aufheben":
                    if st.button("Mangel als behoben markieren"):
                      conn.execute(
                          """
                                                UPDATE fire_facilities 
                                                SET has_defect = 0, defect_description = NULL 
                                                WHERE id = ?
                                            """,
                          (clicked_fac["id"],),
                      )
                      conn.commit()
                      st.success("Mangel behoben!")
                      st.rerun()

                else:
                  st.info(
                      "Klicke auf einen vorhandenen Farbpunkt, um ihn zu"
                      " kontrollieren."
                  )
              else:
                st.info("Klicke auf einen Punkt auf dem Plan.")

      with tab_check:
        st.markdown("### TRVB 117 O Prüfpunkte kontrollieren")

        fac_defects = conn.execute(
            """
                SELECT DISTINCT category FROM fire_facilities 
                WHERE property_id = ? AND (has_defect = 1 OR (next_inspection != 'DAUERHAFT' AND next_inspection < ?))
            """,
            (selected_prop_id, date.today().strftime("%Y-%m-%d")),
        ).fetchall()
        defective_categories = {f["category"] for f in fac_defects}

        templates = conn.execute(
            "SELECT * FROM checklist_templates WHERE interval = ?", (interval,)
        ).fetchall()

        results = {}
        remarks = {}

        for t in templates:
          cat = t["facility_category"]
          has_plan_defect = (
              cat in defective_categories
              or (cat == "TUEREN" and "TUER" in defective_categories)
              or (cat == "LOESCHER" and "LOESCHER" in defective_categories)
          )

          st.markdown(
              f"**{cat}** | *{t['trvb_ref']}*: {t['check_item']}"
          )
          st.caption(f"Anweisung: {t['instruction']}")

          default_val = "MANGEL" if has_plan_defect else "OK"
          default_idx = ["OK", "MANGEL", "N/A"].index(default_val)

          r1, r2 = st.columns([1, 3])
          results[t["id"]] = r1.radio(
              "Ergebnis",
              ["OK", "MANGEL", "N/A"],
              index=default_idx,
              key=f"chk_{run_id}_{t['id']}",
              horizontal=True,
          )

          auto_rem = (
              "Mangel / Prüffrist am Gerät vorhanden (siehe Plan & Liste)"
              if has_plan_defect
              else ""
          )
          remarks[t["id"]] = r2.text_input(
              "Bemerkung", value=auto_rem, key=f"rem_{run_id}_{t['id']}"
          )
          st.divider()

        if st.button("💾 Checklisten-Ergebnisse speichern"):
          cur = conn.cursor()
          cur.execute(
              "DELETE FROM inspection_results WHERE inspection_run_id = ?",
              (run_id,),
          )
          for t_id, res in results.items():
            cur.execute(
                """
                            INSERT INTO inspection_results (inspection_run_id, checklist_template_id, result_status, remark)
                            VALUES (?, ?, ?, ?)
                        """,
                (run_id, t_id, res, remarks.get(t_id, "")),
            )
          conn.commit()
          st.success("Gespeichert! Gehe zu Tab 3 für die Unterschriften.")

      with tab_sign:
        st.markdown("### Digitale Unterschrift & Abschluss")
        col_sig1, col_sig2 = st.columns(2)

        with col_sig1:
          st.markdown("##### 1. Unterschrift Brandschutzbeauftragter (BSB)")
          canvas_bsb = st_canvas(
              fill_color="rgba(255, 255, 255, 0)",
              stroke_width=2,
              stroke_color="#000000",
              background_color="#F8FAFC",
              height=140,
              width=320,
              drawing_mode="freedraw",
              key=f"canvas_bsb_{run_id}",
          )

        with col_sig2:
          st.markdown("##### 2. Kenntnisnahme Betriebsleitung / Eigentümer")
          canvas_mgmt = st_canvas(
              fill_color="rgba(255, 255, 255, 0)",
              stroke_width=2,
              stroke_color="#000000",
              background_color="#F8FAFC",
              height=140,
              width=320,
              drawing_mode="freedraw",
              key=f"canvas_mgmt_{run_id}",
          )

        st.divider()

        if st.button(
            "🏁 Begehung abschließen & Protokoll finalisieren", type="primary"
        ):
          cur = conn.cursor()
          cur.execute(
              "DELETE FROM inspection_results WHERE inspection_run_id = ?",
              (run_id,),
          )
          for t_id, res in results.items():
            cur.execute(
                """
                            INSERT INTO inspection_results (inspection_run_id, checklist_template_id, result_status, remark)
                            VALUES (?, ?, ?, ?)
                        """,
                (run_id, t_id, res, remarks.get(t_id, "")),
            )

          sig_bsb_path = None
          if (
              canvas_bsb.image_data is not None
              and np.any(canvas_bsb.image_data[:, :, 3] > 0)
          ):
            img_bsb = Image.fromarray(canvas_bsb.image_data.astype("uint8"))
            sig_bsb_filename = f"sig_bsb_{run_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            sig_bsb_path = os.path.join(UPLOAD_DIR, sig_bsb_filename)
            img_bsb.save(sig_bsb_path)

          sig_mgmt_path = None
          if (
              canvas_mgmt.image_data is not None
              and np.any(canvas_mgmt.image_data[:, :, 3] > 0)
          ):
            img_mgmt = Image.fromarray(canvas_mgmt.image_data.astype("uint8"))
            sig_mgmt_filename = f"sig_mgmt_{run_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            sig_mgmt_path = os.path.join(UPLOAD_DIR, sig_mgmt_filename)
            img_mgmt.save(sig_mgmt_path)

          cur.execute(
              """
                        UPDATE inspection_runs 
                        SET status = 'ABGESCHLOSSEN', sig_bsb_path = ?, sig_mgmt_path = ? 
                        WHERE id = ?
                    """,
              (sig_bsb_path, sig_mgmt_path, run_id),
          )

          conn.commit()
          st.session_state["finished_run_id"] = run_id
          st.session_state["active_run_id"] = None
          st.success("Begehung erfolgreich abgeschlossen und archiviert!")
          st.rerun()

    if (
        "finished_run_id" in st.session_state
        and st.session_state["finished_run_id"]
    ):
      fin_id = st.session_state["finished_run_id"]
      pdf_bytes = generate_combined_pdf(fin_id)
      st.download_button(
          label="📄 Gesamtes Begehungsprotokoll (inkl. Planpunkten) als PDF"
          " herunterladen",
          data=pdf_bytes,
          file_name=f"Begehungsprotokoll_Prueflauf_{fin_id}.pdf",
          mime="application/pdf",
      )


# ----------------------------------------------------
# 3. LIVE-KAMERA QR-SCANNER (ENTFERNT UM ABSTÜRZE ZU VERHINDERN)
# ----------------------------------------------------
elif menu == "📷 Live-Kamera QR-Scanner":
  st.subheader("📷 Live-Kamera QR-Code-Scanner")
  st.info(
      "Der Live-Kamera-Scan wurde für den Web-Betrieb aufgrund von"
      " Cloud-Schnittstellen deaktiviert. Nutze für Begehungen die"
      " 'Aktive Begehung & Planprüfung' oder das Anlagenkataster."
  )


# ----------------------------------------------------
# 4. ANLAGENKATASTER & FRISTEN
# ----------------------------------------------------
elif menu == "Anlagenkataster & Fristen":
  st.subheader("Anlagenkataster & Prüffristen")

  properties = conn.execute("SELECT * FROM properties").fetchall()
  if not properties:
    st.warning("Bitte zuerst ein Objekt anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox("Objekt wählen", list(prop_dict.keys()))
    selected_prop_id = prop_dict[selected_prop_label]

    facilities = conn.execute(
        """
            SELECT f.*, fp.name as plan_name 
            FROM fire_facilities f
            LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
            WHERE f.property_id = ?
            ORDER BY f.next_inspection ASC
        """,
        (selected_prop_id,),
    ).fetchall()

    today = date.today()
    warn_threshold = today + timedelta(days=60)

    overdue_count = 0
    due_soon_count = 0
    ok_count = 0

    for f in facilities:
      if f["has_defect"]:
        overdue_count += 1
      elif f["next_inspection"] in ["DAUERHAFT", "None", ""]:
        ok_count += 1
      else:
        try:
          next_dt = datetime.strptime(
              f["next_inspection"], "%Y-%m-%d"
          ).date()
          if next_dt < today:
            overdue_count += 1
          elif next_dt <= warn_threshold:
            due_soon_count += 1
          else:
            ok_count += 1
        except ValueError:
          ok_count += 1

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Gesamte Einrichtungen", len(facilities))
    m2.metric("🔴 Mangel / Überfällig", overdue_count)
    m3.metric("🟡 Fällig in < 60 Tagen", due_soon_count)
    m4.metric("🟢 Prüffrist i.O. / Dauerhaft", ok_count)

    st.divider()

    for f in facilities:
      next_dt = None
      if f["next_inspection"] not in ["DAUERHAFT", "None", ""]:
        try:
          next_dt = datetime.strptime(
              f["next_inspection"], "%Y-%m-%d"
          ).date()
        except ValueError:
          pass

      if f["has_defect"]:
        status_badge = f"🔴 **MANGEL:** {f['defect_description'] or 'vorhanden'}"
      elif next_dt and next_dt < today:
        status_badge = "🔴 **ÜBERFÄLLIG**"
      elif next_dt and next_dt <= warn_threshold:
        status_badge = "🟡 **PRÜFUNG FÄLLIG**"
      elif f["next_inspection"] == "DAUERHAFT":
        status_badge = "🟢 **DAUERHAFT GÜLTIG**"
      else:
        status_badge = "🟢 **GÜLTIG**"

      plan_txt = (
          f"📍 Plan: {f['plan_name']}"
          if f["pin_x"]
          else "⚠️ Nicht am Plan verortet"
      )

      with st.expander(
          f"{status_badge} | {f['identifier']} - {f['device_type'] or f['category']} ({f['location_desc']}) | {plan_txt}"
      ):
        c1, c2, c3 = st.columns([2, 2, 2])
        c1.write(f"**Kategorie:** `{f['category']}`")
        c1.write(f"**Standort:** {f['location_desc']}")
        c2.write(f"**Letzte Prüfung:** {f['last_inspection'] or '-'}")
        c2.write(
            f"**Nächste Fälligkeit:**"
            f" `{f['next_inspection'] if f['next_inspection'] != 'DAUERHAFT' else 'Keine (Dauerhaft)'}`"
        )
        c3.write(
            f"**Mangel:** {f['defect_description'] if f['has_defect'] else 'Keiner'}"
        )

        btn_c1, btn_c2 = st.columns(2)
        if f["inspection_interval_months"] > 0:
          if btn_c1.button(
              "🔄 Prüffrist um 2 Jahre verlängern", key=f"cat_r_{f['id']}"
          ):
            new_l = date.today()
            new_n = new_l.replace(year=new_l.year + 2)
            conn.execute(
                """
                            UPDATE fire_facilities 
                            SET last_inspection = ?, next_inspection = ?, has_defect = 0 
                            WHERE id = ?
                        """,
                (
                    new_l.strftime("%Y-%m-%d"),
                    new_n.strftime("%Y-%m-%d"),
                    f["id"],
                ),
            )
            conn.commit()
            st.rerun()

        if btn_c2.button(
            "🗑️ Diese Einrichtung löschen", key=f"del_fac_kat_{f['id']}"
        ):
          conn.execute(
              "DELETE FROM fire_facilities WHERE id = ?", (f["id"],)
          )
          conn.commit()
          st.warning(f"Einrichtung '{f['identifier']}' gelöscht!")
          st.rerun()


# ----------------------------------------------------
# 5. QR-CODE ETIKETTENDRUCK
# ----------------------------------------------------
elif menu == "🏷️ QR-Code Etikettendruck":
  st.subheader("🏷️ QR-Code Etikettendruck für Brandschutzeinrichtungen")
  st.caption(
      "Erstelle druckfertige DIN-A4-Aufkleberbögen mit QR-Codes für"
      " Feuerlöscher, Brandschutztüren, Wandhydranten und Brandmelder zur"
      " eindeutigen Kennzeichnung vor Ort."
  )

  properties = conn.execute("SELECT * FROM properties").fetchall()
  if not properties:
    st.warning("Bitte zuerst ein Objekt unter 'Pläne & Objekte' anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox(
        "Objekt auswählen", list(prop_dict.keys()), key="qr_prop"
    )
    selected_prop_id = prop_dict[selected_prop_label]

    facilities = conn.execute(
        """
            SELECT f.*, fp.name as plan_name 
            FROM fire_facilities f
            LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
            WHERE f.property_id = ?
            ORDER BY f.category, f.identifier ASC
        """,
        (selected_prop_id,),
    ).fetchall()

    if not facilities:
      st.info("Noch keine Brandschutzeinrichtungen für dieses Objekt erfasst.")
    else:
      st.markdown(f"Gefundene Einrichtungen: **{len(facilities)} Stück**")

      col_q1, col_q2 = st.columns([1, 1])
      with col_q1:
        cat_filter = st.multiselect(
            "Nach Kategorie filtern (z.B. nur Feuerlöscher):",
            options=sorted(list({f["category"] for f in facilities})),
            default=[],
        )

      filtered_facs = [
          f
          for f in facilities
          if not cat_filter or f["category"] in cat_filter
      ]

      col_all1, col_all2 = st.columns([2, 8])
      select_all = col_all1.checkbox("Alle auswählen", value=True)

      selected_ids = []
      st.markdown("#### Vorschau & Auswahl der Etiketten:")

      for f in filtered_facs:
        c_chk, c_txt = st.columns([0.5, 9.5])
        is_sel = c_chk.checkbox(
            "", value=select_all, key=f"qr_sel_{f['id']}"
        )
        if is_sel:
          selected_ids.append(f["id"])

        plan_txt = f" | {f['plan_name']}" if f["plan_name"] else ""
        c_txt.write(
            f"**{f['identifier']}** – `{f['category']}` | {f['device_type'] or '-'} |"
            f" Standort: {f['location_desc']}{plan_txt}"
        )

      st.divider()

      if selected_ids:
        pdf_labels = generate_qr_labels_pdf(selected_prop_id, selected_ids)
        st.download_button(
            label=(
                f"🏷️ A4-Etikettenbogen mit QR-Codes als PDF herunterladen"
                f" ({len(selected_ids)} Etiketten)"
            ),
            data=pdf_labels,
            file_name=f"QR_Etiketten_{selected_prop_label.split(' ')[0]}_{date.today()}.pdf",
            mime="application/pdf",
            type="primary",
        )
      else:
        st.warning("Bitte wähle mindestens ein Gerät für den Druckbogen aus.")


# ----------------------------------------------------
# 6. HANDWERKER- & MÄNGELAUFTRAG
# ----------------------------------------------------
elif menu == "🛠️ Handwerker- & Mängelauftrag":
  st.subheader("🛠️ Handwerker- & Instandsetzungsauftrag")
  st.caption(
      "Erstelle kompakte Mängellisten mit Planausschnitten und Fotos für"
      " Haustechniker, Schlosser, Elektriker oder Prüfdienste."
  )

  properties = conn.execute("SELECT * FROM properties").fetchall()
  if not properties:
    st.warning("Bitte zuerst ein Objekt unter 'Pläne & Objekte' anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox(
        "Objekt auswählen", list(prop_dict.keys()), key="hw_prop"
    )
    selected_prop_id = prop_dict[selected_prop_label]

    defects = conn.execute(
        """
            SELECT f.*, fp.name as plan_name 
            FROM fire_facilities f
            LEFT JOIN floor_plans fp ON f.floor_plan_id = fp.id
            WHERE f.property_id = ? AND f.has_defect = 1
            ORDER BY f.floor_plan_id, f.id ASC
        """,
        (selected_prop_id,),
    ).fetchall()

    if not defects:
      st.success("🎉 Keine offenen Mängel für dieses Objekt vorhanden!")
    else:
      st.info(f"Es liegen **{len(defects)} offene Mängel** vor.")

      col_flt1, col_flt2 = st.columns([1, 1])
      with col_flt1:
        cat_filter = st.multiselect(
            "Nach Gewerk / Kategorie filtern (optional):",
            options=sorted(list({d["category"] for d in defects})),
            default=[],
        )

      filtered_defects = [
          d
          for d in defects
          if not cat_filter or d["category"] in cat_filter
      ]

      st.markdown("#### Zu beauftragende Mängel auswählen:")
      selected_defect_ids = []

      for d in filtered_defects:
        c_chk, c_info = st.columns([0.5, 9.5])
        is_sel = c_chk.checkbox("", value=True, key=f"sel_def_{d['id']}")
        if is_sel:
          selected_defect_ids.append(d["id"])

        sev_col = (
            ":red[KRITISCH]"
            if d["defect_severity"] == "KRITISCH"
            else ":orange[MITTEL]"
        )
        plan_str = f"Plan: {d['plan_name']}" if d["plan_name"] else "Ohne Plan"

        with c_info.expander(
            f"🔴 {d['identifier']} ({d['category']}) –"
            f" {d['defect_description']} | {sev_col} | {plan_str}"
        ):
          ci1, ci2 = st.columns([3, 1])
          ci1.write(f"**Montageort:** {d['location_desc']}")
          ci1.write(f"**Typ:** {d['device_type'] or '-'}")
          ci1.write(f"**Frist bis:** `{d['defect_due_date'] or '-'}`")

          if d["defect_photo_path"] and os.path.exists(d["defect_photo_path"]):
            ci2.image(
                d["defect_photo_path"],
                caption="Mängelfoto",
                use_container_width=True,
            )

          if st.button("✅ Mangel als behoben markieren", key=f"fix_{d['id']}"):
            conn.execute(
                """
                            UPDATE fire_facilities 
                            SET has_defect = 0, defect_description = NULL 
                            WHERE id = ?
                        """,
                (d["id"],),
            )
            conn.commit()
            st.success("Mangel als erledigt verbucht!")
            st.rerun()

      st.divider()

      if selected_defect_ids:
        pdf_hw = generate_handwerker_pdf(selected_prop_id, selected_defect_ids)
        st.download_button(
            label=(
                f"📄 Handwerker-Auftrag als PDF herunterladen"
                f" ({len(selected_defect_ids)} Mängel)"
            ),
            data=pdf_hw,
            file_name=(
                f"Handwerker_Auftrag_{selected_prop_label}_{date.today()}.pdf"
            ),
            mime="application/pdf",
            type="primary",
        )
      else:
        st.warning(
            "Bitte wähle mindestens einen Mangel aus, um den Auftrag zu"
            " generieren."
        )


# ----------------------------------------------------
# 7. EREIGNISJOURNAL (TRVB 117 O)
# ----------------------------------------------------
elif menu == "🚨 Ereignisjournal (TRVB 117 O)":
  st.subheader("🚨 Brandschutzbuch – Ereignisjournal (TRVB 117 O)")
  st.caption(
      "Lückenlose Dokumentation von Brandmelderauslösungen (Täuschungs-,"
      " Fehl- & Echtalarme), Evakuierungsübungen, Brandschutzunterweisungen und"
      " Behördenterminen."
  )

  properties = conn.execute("SELECT * FROM properties").fetchall()
  if not properties:
    st.warning("Bitte zuerst ein Objekt unter 'Pläne & Objekte' anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox(
        "Objekt auswählen", list(prop_dict.keys()), key="journal_prop"
    )
    selected_prop_id = prop_dict[selected_prop_label]

    tab_j_list, tab_j_add = st.tabs(
        ["📖 1. Dokumentierte Ereignisse & Export", "➕ 2. Neues Ereignis erfassen"]
    )

    with tab_j_list:
      events = conn.execute(
          """
                SELECT * FROM journal_events 
                WHERE property_id = ? 
                ORDER BY event_date DESC, id DESC
            """,
          (selected_prop_id,),
      ).fetchall()

      alarm_count = sum(1 for e in events if e["event_type"] == "ALARM")
      false_alarm_count = sum(
          1 for e in events if e["sub_type"] in ["TAEUSCHUNG", "FEHLALARM"]
      )
      exercise_count = sum(1 for e in events if e["event_type"] == "UEBUNG")
      training_count = sum(1 for e in events if e["event_type"] == "SCHULUNG")

      jm1, jm2, jm3, jm4 = st.columns(4)
      jm1.metric("Gesamte Einträge", len(events))
      jm2.metric("🚨 Alarme gesamt", alarm_count)
      jm3.metric("⚠️ Täuschungs-/Fehlalarme", false_alarm_count)
      jm4.metric("🏃 Übungen & Schulungen", exercise_count + training_count)

      st.divider()

      col_hdr, col_exp = st.columns([7, 3])
      with col_hdr:
        st.markdown("#### Chronologischer Journalverlauf")
      with col_exp:
        if events:
          j_pdf = generate_journal_pdf(selected_prop_id)
          st.download_button(
              label="📄 Ereignisjournal als PDF exportieren",
              data=j_pdf,
              file_name=(
                  f"Brandschutzbuch_Ereignisjournal_{selected_prop_label}_{date.today()}.pdf"
              ),
              mime="application/pdf",
              type="primary",
          )

      if not events:
        st.info("Noch keine Ereignisse in diesem Brandschutzbuch dokumentiert.")
      else:
        for ev in events:
          badge_icon = "🚨"
          if ev["event_type"] == "UEBUNG":
            badge_icon = "🏃"
          elif ev["event_type"] == "SCHULUNG":
            badge_icon = "🎓"
          elif ev["event_type"] == "BEHOERDE":
            badge_icon = "🏛️"
          elif ev["event_type"] == "HEISSARBEIT":
            badge_icon = "🔥"

          sub_badge = f" | {ev['sub_type']}" if ev["sub_type"] else ""
          fw_badge = (
              " | **🚒 FEUERWEHREINSATZ**" if ev["fire_brigade_deployed"] else ""
          )

          with st.expander(
              f"{badge_icon} {ev['event_date']} – [{ev['event_type']}{sub_badge}]"
              f" {ev['title']}{fw_badge}"
          ):
            c1, c2 = st.columns([3, 1])
            c1.write(f"**Beschreibung / Hergang:** {ev['details'] or '-'}")
            if ev["detector_group"]:
              c1.write(f"**Ausgelöste Meldergruppe:** `{ev['detector_group']}`")
            if ev["instructor_name"]:
              c1.write(f"**Verantwortlich / Prüfer:** {ev['instructor_name']}")

            if ev["participants_count"]:
              c2.metric("Teilnehmer", ev["participants_count"])
            if ev["duration_minutes"]:
              c2.metric("Dauer / Zeit", f"{ev['duration_minutes']} min")

            if st.button("🗑️ Eintrag löschen", key=f"del_ev_{ev['id']}"):
              conn.execute(
                  "DELETE FROM journal_events WHERE id = ?", (ev["id"],)
              )
              conn.commit()
              st.success("Eintrag gelöscht!")
              st.rerun()

    with tab_j_add:
      st.markdown("#### Neues Ereignis im Brandschutzbuch dokumentieren")

      ev_type = st.selectbox(
          "Ereignis-Kategorie*",
          [
              "ALARM",
              "UEBUNG",
              "SCHULUNG",
              "BEHOERDE",
              "HEISSARBEIT",
              "SONSTIGES",
          ],
          format_func=lambda x: {
              "ALARM": (
                  "🚨 Brandmelderauslösung / Brandfall (Echt-, Täuschungs-,"
                  " Fehlalarm)"
              ),
              "UEBUNG": "🏃 Räumungs- / Evakuierungs- / Brandübung",
              "SCHULUNG": (
                  "🎓 Brandschutzunterweisung / Ausbildung / Löschtraining"
              ),
              "BEHOERDE": (
                  "🏛️ Behördentermin (Feuerpolizeiliche Überprüfung / Beschau)"
              ),
              "HEISSARBEIT": (
                  "🔥 Freigabeschein für Heißarbeiten (Schweißen, Trennen)"
              ),
              "SONSTIGES": "📝 Sonstiges Vorkommnis / Reparatur",
          }[x],
      )

      with st.form("add_event_form", clear_on_submit=True):
        col_f1, col_f2 = st.columns(2)
        ev_date = col_f1.date_input("Ereignisdatum*", value=date.today())
        ev_time = col_f2.text_input("Uhrzeit (optional)", placeholder="z.B. 14:25")

        ev_sub = None
        ev_det_group = None
        ev_fw = False
        ev_part_cnt = None
        ev_dur = None
        ev_inst = None

        if ev_type == "ALARM":
          st.markdown("##### Alarmdetails")
          a_c1, a_c2 = st.columns(2)
          ev_sub = a_c1.selectbox(
              "Alarm-Klassifizierung*",
              [
                  "TAEUSCHUNG (Dampf, Staub, Bauarbeiten)",
                  "FEHLALARM (Technischer Defekt / Sensorstörung)",
                  "ECHTALARM (Entstehungsbrand / Brandereignis)",
                  "PROBEALARM (Revision / Wartung)",
              ],
          )
          ev_det_group = a_c2.text_input(
              "Meldergruppe / Melderlinie / Bereich",
              placeholder="z.B. Gruppe 12 / Melder 03 (Lagerhalle)",
          )
          ev_fw = st.checkbox(
              "Feuerwehr alarmiert / ausgerückt", value=False
          )

        elif ev_type == "UEBUNG":
          st.markdown("##### Übungsdetails")
          u_c1, u_c2 = st.columns(2)
          ev_sub = u_c1.selectbox(
              "Übungsart*",
              [
                  "EVAKUIERUNG (Räumungsübung)",
                  "STABSUEBUNG (Brandschutzteam)",
                  "EINSATZUEBUNG (mit Feuerwehr)",
              ],
          )
          ev_dur = u_c2.number_input(
              "Räumungszeit / Dauer (Minuten)", min_value=1, value=4
          )
          ev_part_cnt = st.number_input(
              "Anzahl evakuierter / beteiligter Personen",
              min_value=1,
              value=25,
          )

        elif ev_type == "SCHULUNG":
          st.markdown("##### Schulungsdetails")
          s_c1, s_c2 = st.columns(2)
          ev_sub = s_c1.selectbox(
              "Schulungsart*",
              [
                  "UNTERWEISUNG (Jährliche Mitarbeiterunterweisung)",
                  "PRAXIS (Praktische Handfeuerlöscher-Übung)",
                  "AUSBILDUNG (Brandschutzwart / BSB Fortbildung)",
              ],
          )
          ev_part_cnt = s_c2.number_input(
              "Teilnehmerzahl", min_value=1, value=15
          )
          ev_inst = st.text_input(
              "Vortragender / BSB", value="Johannes Probst"
          )

        elif ev_type == "BEHOERDE":
          ev_sub = st.selectbox(
              "Behördenverfahren*",
              [
                  "Feuerpolizeiliche Überprüfung (§ Beschau)",
                  "Arbeitsinspektorat-Überprüfung",
                  "Sachverständigen-Abnahme / Gutachten",
              ],
          )
          ev_inst = st.text_input(
              "Prüforgan / Behörde / Sachverständiger",
              placeholder="z.B. Bezirkshauptmannschaft / Brandverhütungsstelle",
          )

        elif ev_type == "HEISSARBEIT":
          ev_sub = "Freigabeschein ausgestellt"
          ev_inst = st.text_input(
              "Ausführende Person / Brandwache",
              placeholder="z.B. Fa. Metallbau / Brandwache durch BSB",
          )

        st.markdown("##### Beschreibung & Maßnahmen")
        title = st.text_input(
            "Kurzbezeichnung / Betreff*",
            placeholder=(
                "z.B. Täuschungsalarm durch Reinigungsarbeiten / Jährliche"
                " Räumungsübung Werk 1"
            ),
        )
        details = st.text_area(
            "Ausführlicher Hergang, Ursachenanalyse & getroffene Maßnahmen",
            placeholder=(
                "z.B. Meldergruppe 12 durch Dampfstrahler ausgelöst. BMZ"
                " rückgestellt, Bereich gelüftet. Mitarbeiter nachgeschult."
            ),
        )

        if st.form_submit_button(
            "Ereignis im Brandschutzbuch speichern", type="primary"
        ):
          if title:
            sub_clean = ev_sub.split(" ")[0] if ev_sub else None

            conn.execute(
                """
                            INSERT INTO journal_events (
                                property_id, event_date, event_time, event_type, sub_type, 
                                title, details, detector_group, fire_brigade_deployed, 
                                participants_count, duration_minutes, instructor_name
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                (
                    selected_prop_id,
                    ev_date.strftime("%Y-%m-%d"),
                    ev_time,
                    ev_type,
                    sub_clean,
                    title,
                    details,
                    ev_det_group,
                    1 if ev_fw else 0,
                    ev_part_cnt,
                    ev_dur,
                    ev_inst,
                ),
            )
            conn.commit()
            st.success("Ereignis erfolgreich im Brandschutzbuch archiviert!")
            st.rerun()
          else:
            st.error("Bitte mindestens einen Betreff/Titel eingeben.")


# ----------------------------------------------------
# 8. BSB-JAHRESBERICHT (TRVB 117 O)
# ----------------------------------------------------
elif menu == "📊 BSB-Jahresbericht (TRVB 117 O)":
  st.subheader("📊 Jahresbericht des Brandschutzbeauftragten (TRVB 117 O)")
  st.caption(
      "Automatisierte Jahreszusammenfassung für die Geschäftsleitung,"
      " Betriebsführung und Behörden über durchgeführte Eigenkontrollen,"
      " Schulungsquoten, Mängelstatistik und BMA-Alarme."
  )

  properties = conn.execute("SELECT * FROM properties").fetchall()
  if not properties:
    st.warning("Bitte zuerst ein Objekt unter 'Pläne & Objekte' anlegen.")
  else:
    prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
    selected_prop_label = st.selectbox(
        "Objekt auswählen", list(prop_dict.keys()), key="ar_prop"
    )
    selected_prop_id = prop_dict[selected_prop_label]

    cur_year = date.today().year
    col_y1, col_y2 = st.columns([1, 3])
    rep_year = col_y1.number_input(
        "Berichtsjahr", min_value=2020, max_value=2050, value=cur_year
    )

    runs_yr = conn.execute(
        """
            SELECT * FROM inspection_runs 
            WHERE property_id = ? AND strftime('%Y', inspection_date) = ?
        """,
        (selected_prop_id, str(rep_year)),
    ).fetchall()

    events_yr = conn.execute(
        """
            SELECT * FROM journal_events 
            WHERE property_id = ? AND strftime('%Y', event_date) = ?
        """,
        (selected_prop_id, str(rep_year)),
    ).fetchall()

    facs_total = conn.execute(
        "SELECT * FROM fire_facilities WHERE property_id = ?",
        (selected_prop_id,),
    ).fetchall()
    facs_defect = [f for f in facs_total if f["has_defect"]]

    alarms_yr = [e for e in events_yr if e["event_type"] == "ALARM"]
    trainings_yr = [e for e in events_yr if e["event_type"] == "SCHULUNG"]
    exercises_yr = [e for e in events_yr if e["event_type"] == "UEBUNG"]

    st.markdown(f"#### Jahres-Kennzahlen {rep_year} im Überblick:")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Begehungen / Prüfläufe", len(runs_yr))
    m2.metric("Schulungen & Übungen", len(trainings_yr) + len(exercises_yr))
    m3.metric("BMA-Alarme gesamt", len(alarms_yr))
    m4.metric("Offene Mängel am Stichtag", len(facs_defect))

    st.divider()

    st.markdown("#### Fazit & Gesamtbeurteilung des Brandschutzbeauftragten:")
    default_fazit = (
        f"Im Berichtsjahr {rep_year} wurden alle vorgeschriebenen"
        " Eigenkontrollen gem. TRVB 117 O ordnungsgemäß durchgeführt. Die"
        " Brandschutzeinrichtungen befinden sich in einem betriebsbereiten"
        " Zustand. Festgestellte Mängel wurden bzw. werden im Rahmen des"
        " Instandsetzungsmanagements zeitnah abgearbeitet."
    )
    custom_summary = st.text_area(
        "Individueller Berichtstext für die Geschäftsleitung:",
        value=default_fazit,
        height=100,
    )

    st.divider()

    ar_pdf = generate_annual_report_pdf(
        selected_prop_id, rep_year, custom_summary
    )
    st.download_button(
        label=(
            f"📄 BSB-Jahresbericht {rep_year} als Management-PDF herunterladen"
        ),
        data=ar_pdf,
        file_name=(
            f"BSB_Jahresbericht_{selected_prop_label.split(' ')[0]}_{rep_year}.pdf"
        ),
        mime="application/pdf",
        type="primary",
    )


# ----------------------------------------------------
# 9. DATEN-EXPORT & DATENSICHERUNG
# ----------------------------------------------------
elif menu == "💾 Daten-Export & Datensicherung":
  st.subheader("💾 Daten-Export, Backup & Wiederherstellung")
  st.caption(
      "Exportiere das gesamte Kataster als Excel-Datei oder erstelle ein"
      " 1-Klick-Backup aller Daten und Pläne."
  )

  tab_exp, tab_bak = st.tabs(
      ["📊 1. Excel-Export (.xlsx)", "📦 2. 1-Klick-Backup & Restore (ZIP)"]
  )

  with tab_exp:
    st.markdown("#### Strukturierter Excel-Export")
    properties = conn.execute("SELECT * FROM properties").fetchall()

    if not properties:
      st.warning("Keine Objekte vorhanden.")
    else:
      prop_dict = {"Alle Objekte (Gesamtexport)": None}
      for p in properties:
        prop_dict[f"{p['name']} ({p['address']})"] = p["id"]

      sel_p_label = st.selectbox(
          "Umfang für Excel-Export auswählen:", list(prop_dict.keys())
      )
      target_pid = prop_dict[sel_p_label]

      excel_data = generate_excel_export(target_pid)
      file_name_suffix = (
          "Gesamtexport"
          if target_pid is None
          else sel_p_label.split(" ")[0].replace("/", "_")
      )

      st.download_button(
          label="📊 Excel-Arbeitsmappe (.xlsx) herunterladen",
          data=excel_data,
          file_name=(
              f"Brandschutz_Export_{file_name_suffix}_{date.today().strftime('%Y%m%d')}.xlsx"
          ),
          mime=(
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
          ),
          type="primary",
      )

  with tab_bak:
    st.markdown("#### 1-Klick-Vollsicherung herunterladen")
    st.write(
        "Sichert die gesamte SQLite-Datenbank (`brandschutz.db`) sowie alle"
        " Grundrisspläne, Mängelfotos und Unterschriften in einem ZIP-Archiv."
    )

    zip_bytes = create_full_backup_zip()
    st.download_button(
        label="📦 Vollständiges Backup (ZIP) herunterladen",
        data=zip_bytes,
        file_name=f"Brandschutz_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
        mime="application/zip",
        type="primary",
    )

    st.divider()
    st.markdown("#### Backup wiederherstellen (Restore)")
    st.warning(
        "⚠️ Achtung: Das Wiederherstellen überschreibt die aktuelle Datenbank"
        " und Pläne mit dem Stand aus der ZIP-Datei!"
    )

    uploaded_zip = st.file_uploader(
        "Backup-ZIP-Datei auswählen", type=["zip"], key="zip_restore"
    )
    if uploaded_zip is not None:
      if st.button("🔄 Datensicherung jetzt wiederherstellen", type="secondary"):
        try:
          restore_backup_from_zip(uploaded_zip.read())
          st.success("Datensicherung erfolgreich wiederhergestellt!")
          st.rerun()
        except Exception as e:
          st.error(f"Fehler bei der Wiederherstellung: {e}")


# ----------------------------------------------------
# 10. BRANDSCHUTZBUCH-HISTORIE
# ----------------------------------------------------
elif menu == "Brandschutzbuch-Historie":
  st.subheader("Dokumentierte Eigenkontrollen (Brandschutzbuch)")
  runs = conn.execute("""
        SELECT r.*, p.name as prop_name
        FROM inspection_runs r
        JOIN properties p ON r.property_id = p.id
        ORDER BY r.inspection_date DESC
    """).fetchall()

  if not runs:
    st.info("Noch keine Eigenkontrollen erfasst.")
  else:
    for r in runs:
      c1, c2, c3 = st.columns([5, 2, 2])
      c1.write(
          f"**{r['inspection_date']}** | {r['prop_name']} | Intervall:"
          f" `{r['interval_scope']}` | Prüfer: {r['inspector_name']} | Status:"
          f" `{r['status']}`"
      )

      pdf_data = generate_combined_pdf(r["id"])
      c2.download_button(
          label="📄 PDF Protokoll",
          data=pdf_data,
          file_name=f"Protokoll_{r['prop_name']}_{r['inspection_date']}.pdf",
          mime="application/pdf",
          key=f"dl_run_{r['id']}",
      )

      if c3.button("🗑️ Löschen", key=f"del_run_hist_{r['id']}"):
        run_id_to_del = r["id"]
        conn.execute(
            "DELETE FROM inspection_results WHERE inspection_run_id = ?",
            (run_id_to_del,),
        )
        conn.execute(
            "DELETE FROM defects WHERE inspection_run_id = ?",
            (run_id_to_del,),
        )
        conn.execute(
            "DELETE FROM inspection_runs WHERE id = ?", (run_id_to_del,)
        )
        conn.commit()
        st.warning("Prüfprotokoll gelöscht!")
        st.rerun()
      st.divider()


# ----------------------------------------------------
# 11. PLÄNE & OBJEKTE VERWALTEN
# ----------------------------------------------------
elif menu == "Pläne & Objekte verwalten":
  tab_p1, tab_p2, tab_p3 = st.tabs([
      "🏢 1. Liegenschaften / Objekte verwalten",
      "🗺️ 2. Geschosspläne hochladen & verwalten",
      "➕ 3. Neues Objekt anlegen",
  ])

  with tab_p1:
    st.subheader("Bestehende Objekte / Liegenschaften")
    properties = conn.execute("SELECT * FROM properties").fetchall()

    if not properties:
      st.info("Keine Objekte angelegt.")
    else:
      for p in properties:
        with st.expander(f"🏢 {p['name']} ({p['address']})"):
          st.write(f"**Zuständiger BSB:** {p['fire_safety_officer']}")

          pl_count = conn.execute(
              "SELECT COUNT(*) FROM floor_plans WHERE property_id = ?",
              (p["id"],),
          ).fetchone()[0]
          fac_count = conn.execute(
              "SELECT COUNT(*) FROM fire_facilities WHERE property_id = ?",
              (p["id"],),
          ).fetchone()[0]
          run_count = conn.execute(
              "SELECT COUNT(*) FROM inspection_runs WHERE property_id = ?",
              (p["id"],),
          ).fetchone()[0]
          ev_count = conn.execute(
              "SELECT COUNT(*) FROM journal_events WHERE property_id = ?",
              (p["id"],),
          ).fetchone()[0]

          st.caption(
              f"Verknüpfte Daten: {pl_count} Pläne | {fac_count} Einrichtungen"
              f" | {run_count} Begehungen | {ev_count} Journal-Einträge"
          )

          st.divider()
          st.markdown("##### ⚠️ Objekt löschen")
          st.warning(
              f"Das Löschen von '{p['name']}' entfernt alle zugehörigen"
              " Pläne, Einrichtungen, Prüfprotokolle und Journal-Einträge"
              " unwiderruflich!"
          )

          chk_del = st.checkbox(
              f"Ich möchte '{p['name']}' und alle verknüpften Daten wirklich"
              " löschen",
              key=f"del_prop_chk_{p['id']}",
          )
          if st.button(
              f"🗑️ Objekt '{p['name']}' endgültig löschen",
              disabled=not chk_del,
              key=f"btn_del_prop_{p['id']}",
              type="secondary",
          ):
            prop_id = p["id"]

            plans_to_del = conn.execute(
                "SELECT image_path FROM floor_plans WHERE property_id = ?",
                (prop_id,),
            ).fetchall()
            for pl in plans_to_del:
              if pl["image_path"] and os.path.exists(pl["image_path"]):
                try:
                  os.remove(pl["image_path"])
                except Exception:
                  pass

            runs_to_del = conn.execute(
                "SELECT id FROM inspection_runs WHERE property_id = ?",
                (prop_id,),
            ).fetchall()
            for r in runs_to_del:
              conn.execute(
                  "DELETE FROM inspection_results WHERE inspection_run_id = ?",
                  (r["id"],),
              )
              conn.execute(
                  "DELETE FROM defects WHERE inspection_run_id = ?",
                  (r["id"],),
              )

            conn.execute(
                "DELETE FROM defects WHERE property_id = ?", (prop_id,)
            )
            conn.execute(
                "DELETE FROM inspection_runs WHERE property_id = ?",
                (prop_id,),
            )
            conn.execute(
                "DELETE FROM fire_facilities WHERE property_id = ?",
                (prop_id,),
            )
            conn.execute(
                "DELETE FROM journal_events WHERE property_id = ?",
                (prop_id,),
            )
            conn.execute(
                "DELETE FROM floor_plans WHERE property_id = ?", (prop_id,)
            )

            conn.execute("DELETE FROM properties WHERE id = ?", (prop_id,))
            conn.commit()
            st.success(f"Objekt '{p['name']}' wurde vollständig gelöscht!")
            st.rerun()

  with tab_p2:
    st.subheader("1. Neuen Plan hochladen (PDF oder Bild)")
    properties = conn.execute("SELECT * FROM properties").fetchall()

    if not properties:
      st.warning("Bitte zuerst ein Objekt anlegen.")
    else:
      prop_dict = {f"{p['name']} ({p['address']})": p["id"] for p in properties}
      selected_prop_for_plan = st.selectbox(
          "Objekt für Plan-Upload auswählen",
          list(prop_dict.keys()),
          key="plan_upload_obj",
      )
      target_prop_id = prop_dict[selected_prop_for_plan]

      with st.form("upload_plan_form", clear_on_submit=True):
        plan_name = st.text_input(
            "Planbezeichnung",
            placeholder="z.B. Kellergeschoss / Erdgeschoss / 1. Obergeschoss",
        )
        plan_file = st.file_uploader(
            "Grundrissplan auswählen (PDF, PNG, JPG)",
            type=["pdf", "png", "jpg", "jpeg"],
        )

        if st.form_submit_button("Plan speichern", type="primary"):
          if plan_name and plan_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_ext = os.path.splitext(plan_file.name)[1].lower()

            if file_ext == ".pdf":
              pdf_bytes = plan_file.read()
              doc = fitz.open(stream=pdf_bytes, filetype="pdf")
              page = doc.load_page(0)
              pix = page.get_pixmap(dpi=200)
              saved_filename = f"plan_{target_prop_id}_{timestamp}.png"
              full_path = os.path.join(UPLOAD_DIR, saved_filename)
              pix.save(full_path)
              doc.close()
            else:
              saved_filename = f"plan_{target_prop_id}_{timestamp}{file_ext}"
              full_path = os.path.join(UPLOAD_DIR, saved_filename)
              with open(full_path, "wb") as f:
                f.write(plan_file.getbuffer())

            conn.execute(
                "INSERT INTO floor_plans (property_id, name, image_path) VALUES"
                " (?, ?, ?)",
                (target_prop_id, plan_name, full_path),
            )
            conn.commit()
            st.success(
                f"Plan '{plan_name}' erfolgreich gespeichert und dem Objekt"
                " zugeordnet!"
            )
            st.rerun()
          else:
            st.error("Bitte Bezeichnung eingeben und Datei auswählen.")

      st.divider()
      st.subheader("2. Vorhandene Pläne anzeigen & löschen")
      existing_plans = conn.execute(
          """
                SELECT fp.*, p.name as prop_name 
                FROM floor_plans fp
                JOIN properties p ON fp.property_id = p.id
                WHERE fp.property_id = ?
            """,
          (target_prop_id,),
      ).fetchall()

      if not existing_plans:
        st.info("Noch keine Pläne für dieses Objekt hinterlegt.")
      else:
        for pl in existing_plans:
          col_pl1, col_pl2 = st.columns([7, 3])
          col_pl1.write(f"🗺️ **{pl['name']}** (Hochgeladen am: {pl['created_at']})")
          if col_pl2.button("🗑️ Plan löschen", key=f"del_plan_{pl['id']}"):
            if pl["image_path"] and os.path.exists(pl["image_path"]):
              try:
                os.remove(pl["image_path"])
              except Exception:
                pass
            conn.execute(
                "UPDATE fire_facilities SET floor_plan_id = NULL, pin_x = NULL,"
                " pin_y = NULL WHERE floor_plan_id = ?",
                (pl["id"],),
            )
            conn.execute("DELETE FROM floor_plans WHERE id = ?", (pl["id"],))
            conn.commit()
            st.warning(f"Plan '{pl['name']}' wurde gelöscht!")
            st.rerun()

  with tab_p3:
    st.subheader("Neues Objekt anlegen")
    with st.form("new_prop"):
      name = st.text_input("Objektname / Liegenschaft")
      address = st.text_input("Adresse")
      bsb = st.text_input(
          "Zuständiger Brandschutzbeauftragter (BSB)", value="Johannes Probst"
      )
      if st.form_submit_button("Objekt anlegen", type="primary"):
        if name and address and bsb:
          conn.execute(
              "INSERT INTO properties (name, address, fire_safety_officer)"
              " VALUES (?, ?, ?)",
              (name, address, bsb),
          )
          conn.commit()
          st.success("Objekt angelegt!")
          st.rerun()
        else:
          st.error("Bitte alle Felder ausfüllen.")

conn.close()
